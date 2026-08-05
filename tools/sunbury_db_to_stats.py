"""Convert a Sunbury CC (Grails/MariaDB) database dump into a stats workbook
that imports straight into BetterStats at /admin/import.

    python tools/sunbury_db_to_stats.py dump.sql --out-dir .

The club's old app stored per-innings batting, bowling and fielding records
going back to 1969/70. This reads the dump with no MariaDB install needed,
loads it into a scratch SQLite database, then rolls the innings up into one
row per player per season per team.

Things worth knowing if you adapt this for another club's dump:

* The import wizard reads ``wb.active`` and offers NO sheet picker
  (``import_ingest._parse_xlsx``), so the sheet to import must be both the
  first sheet AND the active one. A read-me tab in front breaks the upload.
* Write an explicit 0 wherever the player actually took part, rather than
  leaving the cell blank. ``row_to_truth`` reconstructs missing not-outs from
  the batting average, and a blank there means an inferred value instead of a
  recorded one. Blank must mean "never recorded", nothing else.
* Overs are cricket notation: 10.3 is ten overs and three balls, not 10.5.
* Counting conventions were validated against the club's own published season
  averages: only a dismissal of 'Not out' counts as a not-out, and Retired
  Hurt counts as an innings.
* The club's stored season averages are stale (they stop at 2019/20), so
  everything here is recomputed from the raw innings records.
"""
import argparse
import csv as _csv
import os
import re
import sqlite3
import sys
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def load_dump(src_path, db_path):
    """Parse the MySQL/MariaDB dump into a fresh SQLite database."""
    SRC = src_path
    if os.path.exists(db_path):
        os.remove(db_path)
    text = open(SRC, encoding="utf-8", errors="replace").read()

    # ── schema: CREATE TABLE blocks, stripped of MySQL-isms ──────────────────────
    con = sqlite3.connect(db_path)
    tables = {}
    for m in re.finditer(r"CREATE TABLE `(\w+)` \((.*?)\n\) ENGINE", text, re.S):
        name, body = m.group(1), m.group(2)
        cols = []
        for line in body.split("\n"):
            line = line.strip().rstrip(",")
            if not line or line.startswith(("PRIMARY KEY", "KEY ", "UNIQUE KEY", "CONSTRAINT")):
                continue
            cm = re.match(r"`(\w+)`\s+(\w+)", line)
            if cm:
                cols.append(cm.group(1))
        tables[name] = cols
        con.execute(f'CREATE TABLE "{name}" ({", ".join(chr(34)+c+chr(34) for c in cols)})')


    def split_tuples(blob):
        """Yield each (...) tuple's values from a MySQL VALUES payload."""
        vals, cur, depth, i, n = [], [], 0, 0, len(blob)
        field, in_str = [], False
        while i < n:
            ch = blob[i]
            if in_str:
                if ch == "\\":
                    nxt = blob[i + 1]
                    field.append({"n": "\n", "t": "\t", "r": "\r", "0": "\0"}.get(nxt, nxt))
                    i += 2
                    continue
                if ch == "'":
                    if i + 1 < n and blob[i + 1] == "'":   # escaped quote
                        field.append("'")
                        i += 2
                        continue
                    in_str = False
                    i += 1
                    continue
                field.append(ch)
                i += 1
                continue
            if ch == "'":
                in_str = True
                field.append("\x00STR")
                i += 1
                continue
            if ch == "(":
                depth += 1
                if depth == 1:
                    cur, field = [], []
                    i += 1
                    continue
            elif ch == ")":
                depth -= 1
                if depth == 0:
                    cur.append("".join(field))
                    vals.append(cur)
                    cur, field = [], []
                    i += 1
                    continue
            elif ch == "," and depth == 1:
                cur.append("".join(field))
                field = []
                i += 1
                continue
            if depth >= 1:
                field.append(ch)
            i += 1
        return vals


    def coerce(raw):
        if raw.startswith("\x00STR"):
            return raw[4:]
        raw = raw.strip()
        if raw.upper() == "NULL":
            return None
        if re.fullmatch(r"-?\d+", raw):
            return int(raw)
        if re.fullmatch(r"-?\d*\.\d+(e-?\d+)?", raw, re.I):
            return float(raw)
        return raw


    total = 0
    for m in re.finditer(r"INSERT INTO `(\w+)` VALUES\s*(.*?);\n", text, re.S):
        tbl, payload = m.group(1), m.group(2)
        if tbl not in tables:
            continue
        rows = [[coerce(v) for v in t] for t in split_tuples(payload)]
        rows = [r for r in rows if len(r) == len(tables[tbl])]
        ph = ",".join("?" * len(tables[tbl]))
        con.executemany(f'INSERT INTO "{tbl}" VALUES ({ph})', rows)
        total += len(rows)

    con.commit()
    con.commit()
    con.close()
    return total


def build_workbook(db_path, out_xlsx, out_csv):
    """Roll the per-innings records up into the import workbook + CSV."""
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row
    q = lambda s, *a: con.execute(s, a).fetchall()

    # ── the one correction we apply: transposed wickets/runs ─────────────────────
    # 19 bowling rows record more wickets than runs conceded, with wickets > 10 —
    # impossible in one innings, and an unambiguous transposition. Left as-is these
    # would produce nonsense season totals (a 67-wicket innings) and nonsense best
    # figures, so they are swapped here and every one is listed on "Data issues".
    SWAPPED = {r["id"] for r in q(
        "select id from bowling_performance where wickets_taken>10 and wickets_taken>runs_conceded")}


    def bowl_figs(row):
        w, r = row["wickets_taken"], row["runs_conceded"]
        return (r, w) if row["id"] in SWAPPED else (w, r)


    def balls_to_overs(balls):
        """Cricket notation: 63 balls -> 10.3 (ten overs and three balls)."""
        return round(balls // 6 + (balls % 6) / 10, 1)


    def fmt_name(first, surname):
        return f"{surname.strip()}, {first.strip()}".strip(", ")


    # ── load context ─────────────────────────────────────────────────────────────
    games = {r["id"]: dict(r) for r in q("""
        select g.id, s.years season, t.name team, gr.name comp, o.name opponent,
               v.name venue, r.name round, gf.match_format format,
               g.result, g.type home_away, g.start_date, g.full_result
        from game g
        join season s on s.id=g.season_id join team t on t.id=g.team_id
        join grade gr on gr.id=g.grade_id join opponent o on o.id=g.opponents_id
        join venue v on v.id=g.venue_id join round r on r.id=g.round_id
        join game_format gf on gf.id=g.game_format_id""")}
    innings_game = {r["id"]: r["game_id"] for r in q("select id, game_id from innings")}
    players = {r["id"]: fmt_name(r["first_name"], r["surname"]) for r in q("select * from player")}


    def gkey(innings_id):
        return innings_game[innings_id]


    # ── accumulate per (player, season, team) and (player, season, comp) ─────────
    def blank():
        return {
            "games": set(), "bat_inns": 0, "runs": 0, "no": 0, "fours": 0, "sixes": 0,
            "f50": 0, "f100": 0, "ducks": 0, "hs": -1, "hs_no": False,
            "bowl_inns": 0, "balls": 0, "maidens": 0, "conceded": 0, "wickets": 0,
            "wides": 0, "nb": 0, "fivew": 0, "best": (-1, 0),
            "ct": 0, "ct_wk": 0, "st": 0, "ro": 0, "field_recs": 0, "flags": set(),
        }


    def collect(team_key):
        """team_key: 'team' (1st XI…) or 'comp' (association grade name)."""
        acc = defaultdict(blank)

        def key(gid, pid):
            g = games[gid]
            return (players[pid], g["season"], g[team_key])

        for r in q("select * from batting_performance"):
            gid = gkey(r["innings_id"])
            a = acc[key(gid, r["player_id"])]
            a["games"].add(gid)
            a["bat_inns"] += 1
            runs, out = r["runs_scored"], r["how_out"]
            not_out = out == "Not out"          # club convention: only 'Not out'
            a["runs"] += runs
            a["no"] += not_out
            a["fours"] += r["fours"]
            a["sixes"] += r["sixes"]
            a["f50"] += 50 <= runs < 100
            a["f100"] += runs >= 100
            a["ducks"] += runs == 0 and not not_out
            if runs > a["hs"] or (runs == a["hs"] and not_out and not a["hs_no"]):
                a["hs"], a["hs_no"] = runs, not_out

        for r in q("select * from bowling_performance"):
            gid = gkey(r["innings_id"])
            a = acc[key(gid, r["player_id"])]
            a["games"].add(gid)
            w, rc = bowl_figs(r)
            if r["id"] in SWAPPED:
                a["flags"].add("swapped bowling figures corrected")
            a["bowl_inns"] += 1
            a["balls"] += r["overs_bowled"] * 6 + (r["extra_balls"] or 0)
            a["maidens"] += r["maidens_bowled"]
            a["conceded"] += rc
            a["wickets"] += w
            a["wides"] += r["wides"]
            a["nb"] += r["no_balls"]
            a["fivew"] += w >= 5
            if (w, -rc) > (a["best"][0], -a["best"][1]):
                a["best"] = (w, rc)

        for r in q("select * from fielding_performance"):
            gid = gkey(r["innings_id"])
            a = acc[key(gid, r["player_id"])]
            a["games"].add(gid)
            a["field_recs"] += 1
            a["ct"] += r["catches"] + r["wicket_keeper_catches"]
            a["ct_wk"] += r["wicket_keeper_catches"]
            a["st"] += r["stumpings"]
            a["ro"] += r["run_outs_assisted"] + r["run_outs_unassisted"]
        return acc


    # ── suspected duplicate entries (same player twice in one innings) ───────────
    dup_rows = []
    for tbl, label in (("batting_performance", "Batting"), ("bowling_performance", "Bowling"),
                       ("fielding_performance", "Fielding")):
        for r in q(f"""select innings_id, player_id, count(*) n,
                         count(distinct {'runs_scored||how_out||batted_at' if tbl=='batting_performance'
                           else 'overs_bowled||runs_conceded||wickets_taken' if tbl=='bowling_performance'
                           else 'catches||stumpings'}) variants
                       from {tbl} group by 1,2 having n>1"""):
            gid = innings_game[r["innings_id"]]
            g = games[gid]
            dup_rows.append({
                "kind": label, "player": players[r["player_id"]], "season": g["season"],
                "team": g["team"], "date": str(g["start_date"])[:10], "opponent": g["opponent"],
                "format": g["format"], "copies": r["n"],
                "verdict": ("Identical rows - almost certainly entered twice" if r["variants"] == 1
                            else "Different figures - may be a genuine 2nd innings (two-day match)"),
            })

    dup_keys = {(d["player"], d["season"], d["team"]) for d in dup_rows}

    # ── build sheets ─────────────────────────────────────────────────────────────
    SEASON_HEADERS = [
        "Player", "Season", "Grade", "Games", "Innings", "Runs", "NO", "HS", "Avg",
        "4s", "6s", "50s", "100s", "Ducks",
        "Overs", "Maidens", "Runs Conceded", "Wickets", "Bowl Avg", "Econ", "Best", "5WI",
        "Wides", "No Balls",
        "Catches", "Catches WK", "Stumpings", "Run Outs", "Review flag",
    ]


    def rows_for(acc, flag_dups=True):
        out = []
        for (name, season, grade), a in acc.items():
            if not a["games"]:
                continue
            outs = a["bat_inns"] - a["no"]
            avg = round(a["runs"] / outs, 2) if outs else None
            bavg = round(a["conceded"] / a["wickets"], 2) if a["wickets"] else None
            econ = round(a["conceded"] / (a["balls"] / 6), 2) if a["balls"] else None
            hs = None
            if a["hs"] >= 0:
                hs = f"{a['hs']}*" if a["hs_no"] else a["hs"]
            best = f"{a['best'][0]}-{a['best'][1]}" if a["best"][0] >= 0 else None
            flags = set(a["flags"])
            if flag_dups and (name, season, grade) in dup_keys:
                flags.add("possible duplicate entry - see Data issues")
            # A zero is only written where the player actually took part in that
            # discipline — there a 0 is a recorded fact. Where they never batted,
            # bowled or fielded the cells stay blank, which BetterStats reads as
            # "not recorded" rather than "none". Not-outs in particular MUST be an
            # explicit 0, or row_to_truth reconstructs them from the average.
            bat = (lambda v: v) if a["bat_inns"] else (lambda v: None)
            bowl = (lambda v: v) if a["bowl_inns"] else (lambda v: None)
            fld = (lambda v: v) if a["field_recs"] else (lambda v: None)
            out.append([
                name, season, grade, len(a["games"]),
                bat(a["bat_inns"]), bat(a["runs"]), bat(a["no"]), hs, avg,
                bat(a["fours"]), bat(a["sixes"]), bat(a["f50"]), bat(a["f100"]), bat(a["ducks"]),
                bowl(balls_to_overs(a["balls"])), bowl(a["maidens"]),
                bowl(a["conceded"]), bowl(a["wickets"]), bavg, econ, best, bowl(a["fivew"]),
                bowl(a["wides"]), bowl(a["nb"]),
                fld(a["ct"]), fld(a["ct_wk"]), fld(a["st"]), fld(a["ro"]),
                "; ".join(sorted(flags)) or None,
            ])
        return sorted(out, key=lambda r: (r[1], r[2], r[0]))


    by_team = rows_for(collect("team"))
    by_comp = rows_for(collect("comp"), flag_dups=False)

    # career totals
    career = defaultdict(blank)
    acc_team = collect("team")
    for (name, season, grade), a in acc_team.items():
        c = career[name]
        c["games"] |= a["games"]
        for k in ("bat_inns", "runs", "no", "fours", "sixes", "f50", "f100", "ducks",
                  "bowl_inns", "balls", "maidens", "conceded", "wickets", "wides", "nb",
                  "fivew", "ct", "ct_wk", "st", "ro"):
            c[k] += a[k]
        if a["hs"] > c["hs"] or (a["hs"] == c["hs"] and a["hs_no"] and not c["hs_no"]):
            c["hs"], c["hs_no"] = a["hs"], a["hs_no"]
        if (a["best"][0], -a["best"][1]) > (c["best"][0], -c["best"][1]):
            c["best"] = a["best"]

    career_rows = []
    for name, a in career.items():
        outs = a["bat_inns"] - a["no"]
        seasons = sorted({s for (n, s, _g) in acc_team if n == name})
        career_rows.append([
            name, len(seasons), f"{seasons[0][:4]}-{seasons[-1][5:]}" if seasons else "",
            len(a["games"]), a["bat_inns"] or None, a["runs"] or None, a["no"] or None,
            (f"{a['hs']}*" if a["hs_no"] else a["hs"]) if a["hs"] >= 0 else None,
            round(a["runs"] / outs, 2) if outs else None,
            a["f50"] or None, a["f100"] or None, a["ducks"] or None,
            balls_to_overs(a["balls"]) or None, a["maidens"] or None, a["conceded"] or None,
            a["wickets"] or None,
            round(a["conceded"] / a["wickets"], 2) if a["wickets"] else None,
            round(a["conceded"] / (a["balls"] / 6), 2) if a["balls"] else None,
            f"{a['best'][0]}-{a['best'][1]}" if a["best"][0] >= 0 else None, a["fivew"] or None,
            a["ct"] or None, a["ct_wk"] or None, a["st"] or None, a["ro"] or None,
        ])
    career_rows.sort(key=lambda r: -(r[5] or 0))

    # match list
    match_rows = []
    for gid, g in sorted(games.items(), key=lambda kv: str(kv[1]["start_date"])):
        inn = q("select innings, extras, opponent_runs, opponent_extras from innings where game_id=?", gid)
        match_rows.append([
            str(g["start_date"])[:10], g["season"], g["team"], g["comp"], g["round"],
            g["format"], g["home_away"], g["opponent"], g["venue"], g["result"],
            "; ".join(str(i["innings"]) for i in inn),
            "; ".join(str(i["opponent_runs"]) for i in inn),
            g["full_result"],
        ])

    # ── write workbook ───────────────────────────────────────────────────────────
    # NOTE: BetterStats' import wizard reads wb.active and offers no sheet picker
    # (import_ingest._parse_xlsx). "Season Stats" must therefore be BOTH the first
    # sheet and the active one, or the upload reads whatever else is in front of it.
    wb = Workbook()
    HEAD = PatternFill("solid", fgColor="16A34A")
    HEADF = Font(bold=True, color="FFFFFF", size=10)
    TITLE = Font(bold=True, size=14)
    SUB = Font(italic=True, color="555555", size=10)
    THIN = Border(bottom=Side(style="thin", color="DDDDDD"))


    def sheet(name, headers, rows, widths=None, freeze="A2"):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for cell in ws[1]:
            cell.fill, cell.font = HEAD, HEADF
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r in rows:
            ws.append(r)
        ws.freeze_panes = freeze
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
        for i, h in enumerate(headers, 1):
            w = (widths or {}).get(h, max(9, min(30, len(str(h)) + 3)))
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 30
        return ws


    W = {"Player": 24, "Season": 11, "Grade": 34, "Best": 8, "Review flag": 46,
         "Runs Conceded": 11, "Catches WK": 11, "Run Outs": 10, "Overs": 9}

    # Sheet 1, and the active one: the import target.
    wb.remove(wb.active)
    sheet("Season Stats", SEASON_HEADERS, by_team, W)

    # Read me
    ws = wb.create_sheet("Read me")
    ws.column_dimensions["A"].width = 118
    lines = [
        ("Sunbury Cricket Club - historical statistics", TITLE),
        (f"Built from the club's MariaDB database dump. {len(games):,} matches, "
         f"{len(players):,} players, 1969/70 to 2023/24.", SUB),
        ("", None),
        ("What's in this file", Font(bold=True, size=11)),
        ("  Season Stats        One row per player, per season, per team. This is the sheet that imports.", None),
        ("  Season by Comp      The same numbers split by association grade instead of by XI.", None),
        ("  Career Totals       All-time figures per player, for reference.", None),
        ("  Matches             Every match in the database, for reference.", None),
        ("  Data issues         Everything worth a second look before you trust the numbers.", None),
        ("", None),
        ("How to import it", Font(bold=True, size=11)),
        ("  1. Go to betterat.cricket/admin/import.", None),
        ("  2. Upload this file. Leave Season Stats as the sheet showing when you save it - the importer reads", None),
        ("     the sheet the workbook was left open on, and there is no sheet picker in the wizard.", None),
        ("  3. The column headings are named so the wizard maps all 28 of them on its own. Check the mapping,", None),
        ("     then match the player names to your roster. The wizard suggests close matches.", None),
        ("  4. If anything looks off, import Sunbury_CC_season_stats.csv instead. It is this same sheet on its", None),
        ("     own, so there is no sheet to get wrong.", None),
        ("", None),
        ("The 'Grade' column holds the XI (1st XI, 2nd XI and so on), because that stays put while the", None),
        ("association's grade names change from year to year. If you would rather import against the real", None),
        ("grade names, use the Season by Comp sheet instead. Don't import both, they are the same runs twice.", None),
        ("", None),
        ("How the numbers were worked out", Font(bold=True, size=11)),
        ("  Everything is added up from the ball-by-ball innings records, not from the club's stored season", None),
        ("  averages. Those stored averages stop at 2019/20, so the last three seasons are missing from them.", None),
        ("  Where the stored averages were up to date, these figures reproduce them exactly.", None),
        ("", None),
        ("  Not out       Only a dismissal of 'Not out' counts. Retired Hurt counts as an innings, matching", None),
        ("                the club's own records.", None),
        ("  Ducks         Out for 0. A score of 0 not out is not counted, which is the one place these figures", None),
        ("                differ on purpose from the club's old 'zeroes' column, which counted both.", None),
        ("  Overs         Cricket notation, so 10.3 means ten overs and three balls.", None),
        ("  Catches       Total catches, keeper's catches included. 'Catches WK' is the keeper's share of them.", None),
        ("  Games         Matches where the player batted, bowled or took a fielding record. Someone who did", None),
        ("                none of the three in a match doesn't show up as having played it.", None),
        ("", None),
        ("One correction was applied", Font(bold=True, size=11)),
        (f"  {len(SWAPPED)} bowling records have wickets and runs the wrong way round (a 6-over spell recorded as", None),
        ("  67 wickets for 1 run). They are swapped back here, and every one is listed on the Data issues", None),
        ("  sheet with its original figures so you can check them. Nothing else was changed.", None),
        ("", None),
        ("  Blank cells mean the club never recorded that number, which is not the same as zero. Balls faced", None),
        ("  and strike rate were never recorded at all, so those columns aren't here.", None),
    ]
    for text, font in lines:
        ws.append([text])
        if font:
            ws.cell(row=ws.max_row, column=1).font = font

    sheet("Season by Comp", SEASON_HEADERS, by_comp, W)
    sheet("Career Totals",
          ["Player", "Seasons", "Span", "Games", "Innings", "Runs", "NO", "HS", "Avg",
           "50s", "100s", "Ducks", "Overs", "Maidens", "Runs Conceded", "Wickets",
           "Bowl Avg", "Econ", "Best", "5WI", "Catches", "Catches WK", "Stumpings", "Run Outs"],
          career_rows, W)
    sheet("Matches",
          ["Date", "Season", "Team", "Competition", "Round", "Format", "H/A", "Opponent",
           "Venue", "Result", "Sunbury score", "Opponent score", "Full result"],
          match_rows, {"Competition": 34, "Opponent": 26, "Venue": 26, "Full result": 40,
                       "Sunbury score": 13, "Opponent score": 13})

    # Data issues
    issue_rows = []
    for r in q("""select bp.id, p.first_name, p.surname, bp.overs_bowled, bp.maidens_bowled,
                         bp.runs_conceded, bp.wickets_taken, i.game_id
                  from bowling_performance bp join player p on p.id=bp.player_id
                  join innings i on i.id=bp.innings_id where bp.id in (%s)"""
               % ",".join(map(str, SWAPPED or {0}))):
        g = games[r["game_id"]]
        issue_rows.append([
            "Wickets/runs swapped", fmt_name(r["first_name"], r["surname"]), g["season"],
            g["team"], str(g["start_date"])[:10], g["opponent"],
            f"{r['overs_bowled']} ov, {r['wickets_taken']} wkts for {r['runs_conceded']} runs",
            f"{r['overs_bowled']} ov, {r['runs_conceded']} wkts for {r['wickets_taken']} runs",
            "Corrected in this file",
        ])
    for d in sorted(dup_rows, key=lambda x: (x["season"], x["player"])):
        issue_rows.append([
            f"{d['kind']} recorded {d['copies']}x", d["player"], d["season"], d["team"],
            d["date"], d["opponent"], f"{d['format']}, {d['copies']} copies of this record", "",
            d["verdict"] + " - left as found",
        ])
    sheet("Data issues",
          ["Issue", "Player", "Season", "Team", "Date", "Opponent",
           "As recorded", "Read as", "What was done"],
          issue_rows,
          {"Issue": 22, "Player": 24, "As recorded": 40, "Read as": 34, "What was done": 52,
           "Opponent": 24})

    wb.active = 0          # the importer reads wb.active; make it Season Stats
    wb.save(out_xlsx)

    # A single-sheet CSV of exactly the same rows, for when a workbook's active
    # sheet can't be relied on (someone opens it, clicks another tab, saves).
    import csv as _csv
    with open(out_csv, "w", newline="", encoding="utf-8-sig") as fh:
        w = _csv.writer(fh)
        w.writerow(SEASON_HEADERS)
        for r in by_team:
            # 2.0 -> 2, so the file reads the way a scorer would write it. Overs
            # keep their decimal (10.3 is ten overs and three balls, not ten point
            # three), and averages keep theirs.
            w.writerow(["" if v is None else
                        int(v) if isinstance(v, float) and v.is_integer() else v
                        for v in r])

    con.close()
    return {
        "season_rows": len(by_team),
        "comp_rows": len(by_comp),
        "career_rows": len(career_rows),
        "matches": len(match_rows),
        "issues": len(issue_rows),
        "corrected": len(SWAPPED),
        "flagged": len(dup_rows),
    }


def main():
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("dump", help="path to the .sql dump")
    ap.add_argument("--out-dir", default=".", help="where to write the files")
    ap.add_argument("--prefix", default="Sunbury_CC", help="output filename prefix")
    ap.add_argument("--keep-db", action="store_true", help="keep the scratch SQLite file")
    args = ap.parse_args()

    if not os.path.exists(args.dump):
        sys.exit(f"no such dump: {args.dump}")
    os.makedirs(args.out_dir, exist_ok=True)
    db_path = os.path.join(args.out_dir, f"{args.prefix}_scratch.sqlite")
    xlsx = os.path.join(args.out_dir, f"{args.prefix}_stats_import.xlsx")
    csv_path = os.path.join(args.out_dir, f"{args.prefix}_season_stats.csv")

    rows = load_dump(args.dump, db_path)
    print(f"loaded {rows:,} rows from {os.path.basename(args.dump)}")
    stats = build_workbook(db_path, xlsx, csv_path)
    if not args.keep_db:
        os.remove(db_path)

    print(f"wrote {xlsx}")
    print(f"wrote {csv_path}")
    print(f"  Season Stats   {stats['season_rows']:,} rows  (the sheet that imports)")
    print(f"  Season by Comp {stats['comp_rows']:,} rows")
    print(f"  Career Totals  {stats['career_rows']:,} rows")
    print(f"  Matches        {stats['matches']:,} rows")
    print(f"  Data issues    {stats['issues']:,} rows "
          f"({stats['corrected']} corrected, {stats['flagged']} flagged)")


if __name__ == "__main__":
    main()
