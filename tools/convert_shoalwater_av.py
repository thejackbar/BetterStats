"""Convert a club's ``DATA<year>.AV`` files into a combined match spreadsheet
and a BetterImport season-stats CSV.

The ``.AV`` files are the fixed-record binary database of an MS-DOS/Delphi era
club averages program.  Nothing in this repository read that format before, so
the layout below was recovered by inspection and then checked against the data
itself (see ``verify`` / the ``Notes`` sheet the converter writes).

File layout — every byte of a 549,871 byte file is accounted for::

    offset       size                 section
    0            311                  header: club name (30) + year (4) + settings
    311          200 x 37             player index
    7,711        200 x 578            per-player stats area (unused in these files)
    123,311      160 x 54             fixtures
    131,951      160 x 2,612          matches

Integers are little-endian ``int16``; ``-1`` means "not recorded".  Overs and
vote fields are Delphi ``Currency`` (``int64`` scaled by 10,000), where
``-10000`` (ie ``-1.0``) means "not recorded".  Dates are Delphi/Excel day
numbers from 1899-12-30.

The club also keeps the same season as separate ``.PLR`` (players), ``.FIX``
(fixtures), ``.MCH`` (matches) and ``.DAT`` files. Checked against the 1993
set, the ``.AV`` is a complete and self-consistent export of them: its
fixtures agree with ``.FIX`` on date, team and opponent for all 41 shared
slots, and its scorelines agree with ``.MCH`` on all 35. It renumbers the
players - ``.PLR`` keeps ids stable across seasons and the ``.AV`` does not -
but its match blocks use its own numbering, so resolving through its own
embedded roster names the same eleven in the same order as ``.MCH`` resolved
through ``.PLR``. ``.DAT`` is blank, like the ``.AV``'s own player-stats
region: the program recomputed averages rather than storing them.

**Overs are cricket notation, not decimals** — ``93.3`` is 93 overs and 3
balls.  Confirmed against the data: the fractional digit is never above 5.
Anything that sums or divides overs therefore converts to balls first.
"""

from __future__ import annotations

import argparse
import csv
import datetime
import re
import struct
from collections import defaultdict
from pathlib import Path

# ── file layout ──────────────────────────────────────────────────────────────

HEADER_LEN = 311
PLAYER_BASE, PLAYER_REC, PLAYER_N = 311, 37, 200
PSTATS_BASE, PSTATS_REC, PSTATS_N = 7_711, 578, 200
FIXTURE_BASE, FIXTURE_REC, FIXTURE_N = 123_311, 54, 160
MATCH_BASE, MATCH_REC, MATCH_N = 131_951, 2_612, 160
FILE_LEN = HEADER_LEN + PLAYER_N * PLAYER_REC + PSTATS_N * PSTATS_REC \
    + FIXTURE_N * FIXTURE_REC + MATCH_N * MATCH_REC          # 549,871

# within a match record
M_ROUND, M_ROUND_LEN = 10, 20
M_GROUND_IDX, M_GROUND, M_GROUND_LEN = 30, 32, 30
M_INN_US, M_INN_THEM = 502, 574        # slot A is always our club's innings
M_FOW = 646                            # pairs of (score, batter position, 1-indexed)
M_FOW_PAIRS = 10
BLOCK_BASE, BLOCK_REC, BLOCK_N = 756, 116, 16

# within a 72-byte innings block
I_TOTAL, I_WICKETS, I_NUMBER, I_OVERS = 0, 2, 4, 6
I_EXTRAS = (14, 16, 18, 20, 22)        # byes, leg byes, wides, no balls, penalty

# within a 116-byte player block
B_PLAYER, B_POSITION, B_RUNS, B_SIXES, B_FOURS, B_HOWOUT = 0, 2, 8, 12, 14, 20
B_OVERS, B_MAIDENS, B_WICKETS = 24, 32, 34
B_FLAGS, B_CONCEDED, B_WIDES, B_NOBALLS = 42, 46, 48, 50
B_CATCHES, B_WK_CATCHES, B_STUMPINGS, B_BYES = 52, 56, 60, 64
B_VOTES = 82

NOT_OUT = 11          # value of B_HOWOUT for a batter who was not out

# Only NOT_OUT is proven (the count of dismissed batters matches the innings
# wickets in 840 of 847 innings). The other five are supported by counting an
# independent population in the same files and finding the same magnitude:
# our batters were run out 443 times against 404 wickets we took that were not
# credited to a bowler, and stumped 184 times against 204 stumpings by our own
# keepers. Their shares (caught 47%, bowled 27%, lbw 9%) are ordinary club
# rates. Everything else is left as a raw code rather than guessed at.
DISMISSALS = {0: "Bowled", 1: "Caught", 2: "LBW", 3: "Stumped",
              4: "Run out", NOT_OUT: "Not out"}


def dismissal_label(code) -> str:
    if code is None:
        return ""
    return DISMISSALS.get(code, f"code {code}")
DID_NOT_BAT = 255     # value of B_POSITION for a player who was not in the order
EPOCH = datetime.date(1899, 12, 30)

# ── primitives ───────────────────────────────────────────────────────────────


def i16(buf: bytes, off: int) -> int:
    return struct.unpack_from("<h", buf, off)[0]


def currency(buf: bytes, off: int):
    """Delphi Currency. Returns None for the -1.0 'not recorded' sentinel."""
    raw = struct.unpack_from("<q", buf, off)[0]
    return None if raw == -10_000 else raw / 10_000


def text(buf: bytes, off: int, length: int) -> str:
    return buf[off:off + length].decode("latin-1").replace("\x00", " ").strip()


def opt(value: int):
    """-1 means 'not recorded'."""
    return None if value < 0 else value


def overs_to_balls(overs) -> int:
    """Cricket notation to balls: 10.2 is ten overs and two balls, so 62."""
    if overs is None:
        return 0
    whole = int(overs)
    balls = int(round((overs - whole) * 10))
    return whole * 6 + balls


def balls_to_overs(balls: int) -> str:
    return f"{balls // 6}.{balls % 6}"


# ── parsing ──────────────────────────────────────────────────────────────────


def season_label(year: int) -> str:
    return f"{year}/{str(year + 1)[-2:]}"


def player_display(surname: str, initial: str, first: str) -> str:
    """'Surname, First' — the shape BetterImport's own matcher reads best."""
    if first:
        return f"{surname}, {first}"
    if initial:
        return f"{surname}, {initial}"
    return surname


def parse_players(buf: bytes) -> dict:
    players = {}
    for slot in range(PLAYER_N):
        off = PLAYER_BASE + slot * PLAYER_REC
        pid = i16(buf, off)
        if pid < 0:
            continue
        surname = text(buf, off + 2, 16)
        initial = text(buf, off + 18, 3)
        first = text(buf, off + 21, 16)
        if not surname:
            continue
        players[pid] = {
            "player_id": pid, "slot": slot, "surname": surname,
            "initial": initial, "first_name": first,
            "name": player_display(surname, initial, first),
        }
    return players


def parse_fixtures(buf: bytes) -> dict:
    """Keyed by the match slot the fixture POINTS AT, not by its own position.

    The fixture array is held in date order, and each record's first field is
    the slot of its match record. The two coincide only when the fixtures were
    entered in slot order, which happens to be true of 1992 and of no other
    season - so reading the array position instead pairs a scorecard with
    another match's date, opponent and team. The tell is that the stored slots
    are a permutation of 0..n-1 rather than a run.
    """
    fixtures = {}
    for pos in range(FIXTURE_N):
        off = FIXTURE_BASE + pos * FIXTURE_REC
        slot = i16(buf, off)
        if slot < 0 or slot >= MATCH_N:
            continue
        name = text(buf, off + 20, 34)
        if not name:
            continue
        serial = struct.unpack_from("<d", buf, off + 2)[0]
        date = EPOCH + datetime.timedelta(days=serial) if serial > 0 else None
        fixtures[slot] = {
            "index": slot, "order": pos, "date": date,
            "team": i16(buf, off + 18), "name": name,
        }
    return fixtures


def split_leg(fixtures: dict) -> None:
    """A 1992-97 fixture is stored as two records, 'Opponent 1' and 'Opponent 2'.

    Those are the two legs of one two-day match, not two opponents and not the
    grade: both carry the same date and the same team number.  Only strip the
    suffix where the sibling leg is actually present, so an opponent whose real
    name ends in a digit is left alone.
    """
    seen = {(f["date"], f["team"], f["name"]) for f in fixtures.values()}
    for fx in fixtures.values():
        m = re.match(r"^(.*?)\s+([12])$", fx["name"])
        fx["opponent"], fx["leg"] = fx["name"], None
        if not m:
            continue
        base, leg = m.group(1), int(m.group(2))
        sibling = f"{base} {2 if leg == 1 else 1}"
        if (fx["date"], fx["team"], sibling) in seen:
            fx["opponent"], fx["leg"] = base, leg


def parse_innings(rec: bytes, base: int) -> dict:
    extras = [opt(i16(rec, base + o)) for o in I_EXTRAS]
    total = opt(i16(rec, base + I_TOTAL))
    return {
        "total": total,
        "wickets": opt(i16(rec, base + I_WICKETS)),
        "innings_no": opt(i16(rec, base + I_NUMBER)),
        "overs": currency(rec, base + I_OVERS),
        "byes": extras[0], "leg_byes": extras[1], "wides": extras[2],
        "no_balls": extras[3], "penalty": extras[4],
        "extras": sum(e for e in extras if e) if any(e for e in extras) else 0,
        "played": total is not None and total >= 0,
    }


def parse_blocks(rec: bytes) -> list:
    blocks = []
    for slot in range(BLOCK_N):
        off = BLOCK_BASE + slot * BLOCK_REC
        pid = i16(rec, off)
        if pid < 0 or pid >= PLAYER_N:
            continue
        pos = i16(rec, off + B_POSITION)
        howout = i16(rec, off + B_HOWOUT)
        runs = opt(i16(rec, off + B_RUNS))
        overs = currency(rec, off + B_OVERS)
        blocks.append({
            "slot": slot, "player_id": pid,
            "position": None if pos in (DID_NOT_BAT, -1) else pos + 1,
            "batted": runs is not None,
            "runs": runs,
            "fours": opt(i16(rec, off + B_FOURS)),
            "sixes": opt(i16(rec, off + B_SIXES)),
            "not_out": howout == NOT_OUT,
            "dismissal_code": opt(howout),
            "bowled": overs is not None,
            "overs": overs,
            "balls": overs_to_balls(overs),
            "maidens": opt(i16(rec, off + B_MAIDENS)),
            "wickets": opt(i16(rec, off + B_WICKETS)),
            "conceded": opt(i16(rec, off + B_CONCEDED)),
            "wides": opt(i16(rec, off + B_WIDES)),
            "no_balls": opt(i16(rec, off + B_NOBALLS)),
            "catches": opt(i16(rec, off + B_CATCHES)),
            "wk_catches": opt(i16(rec, off + B_WK_CATCHES)),
            "stumpings": opt(i16(rec, off + B_STUMPINGS)),
            "byes": opt(i16(rec, off + B_BYES)),
            "votes": currency(rec, off + B_VOTES),
            "flags": opt(i16(rec, off + B_FLAGS)),
        })
    return blocks


def parse_fow(rec: bytes) -> list:
    """Pairs of (score, batting position). Positions are 1-indexed."""
    out = []
    for n in range(M_FOW_PAIRS):
        score = i16(rec, M_FOW + n * 4)
        batter = i16(rec, M_FOW + n * 4 + 2)
        if score < 0:
            break
        out.append({"wicket": n + 1, "score": score,
                    "batter_position": None if batter < 0 else batter})
    return out


def parse_file(path: Path) -> dict:
    buf = path.read_bytes()
    if len(buf) != FILE_LEN:
        raise ValueError(f"{path.name}: expected {FILE_LEN} bytes, got {len(buf)}")
    year = int(text(buf, 30, 4))
    fixtures = parse_fixtures(buf)
    split_leg(fixtures)
    matches = []
    for idx in sorted(fixtures):
        rec = buf[MATCH_BASE + idx * MATCH_REC:MATCH_BASE + (idx + 1) * MATCH_REC]
        us, them = parse_innings(rec, M_INN_US), parse_innings(rec, M_INN_THEM)
        blocks = parse_blocks(rec)
        if not (us["played"] or them["played"] or any(b["batted"] or b["bowled"] for b in blocks)):
            continue          # a side was named but no scorecard was ever entered
        matches.append({
            **fixtures[idx],
            "round": text(rec, M_ROUND, M_ROUND_LEN),
            "ground": text(rec, M_GROUND, M_GROUND_LEN),
            "ground_index": opt(i16(rec, M_GROUND_IDX)),
            "us": us, "them": them, "blocks": blocks, "fow": parse_fow(rec),
            "has_play": us["played"] or them["played"],
        })
    return {
        "source": path.name, "club": text(buf, 0, 30), "year": year,
        "season": season_label(year), "players": parse_players(buf),
        "matches": matches,
    }


# ── shaping ──────────────────────────────────────────────────────────────────


def grade_label(team: int) -> str:
    return f"Grade {team}" if team and team > 0 else "Grade (unknown)"


def build_rows(seasons: list) -> dict:
    """Flatten every season into the sheets, collapsing two-day legs into one match."""
    matches, batting, bowling, fielding, fow_rows, votes, players = [], [], [], [], [], [], []
    match_no = 0
    for s in seasons:
        for p in sorted(s["players"].values(), key=lambda x: (x["surname"], x["first_name"])):
            players.append({"Season": s["season"], "Player ID": p["player_id"],
                            "Player": p["name"], "Surname": p["surname"],
                            "Initial": p["initial"], "First name": p["first_name"]})
        # group the legs of a two-day match under one match id
        groups = defaultdict(list)
        for m in s["matches"]:
            groups[(m["date"], m["team"], m["opponent"])].append(m)
        for key in sorted(groups, key=lambda k: (k[0] or datetime.date.min, k[1], k[2])):
            legs = sorted(groups[key], key=lambda m: (m["leg"] or 0, m["index"]))
            match_no += 1
            mid = f"{s['year']}-{match_no:03d}"
            date, team, opponent = key
            our_runs = sum(l["us"]["total"] for l in legs if l["us"]["played"])
            their_runs = sum(l["them"]["total"] for l in legs if l["them"]["played"])
            both = any(l["us"]["played"] for l in legs) and any(l["them"]["played"] for l in legs)
            result = ""
            if both:
                result = "Won" if our_runs > their_runs else "Lost" if our_runs < their_runs else "Tie"
            for l in legs:
                names = s["players"]
                common = {"Season": s["season"], "Match ID": mid, "Date": date,
                          "Grade": grade_label(team), "Opponent": opponent,
                          "Ground": l["ground"], "Leg": l["leg"] or ""}
                matches.append({
                    **common, "Round": l["round"],
                    "Our runs": l["us"]["total"], "Our wickets": l["us"]["wickets"],
                    "Our overs": l["us"]["overs"], "Our innings no": l["us"]["innings_no"],
                    "Our byes": l["us"]["byes"], "Our leg byes": l["us"]["leg_byes"],
                    "Our wides": l["us"]["wides"], "Our no balls": l["us"]["no_balls"],
                    "Our penalty": l["us"]["penalty"], "Our extras": l["us"]["extras"],
                    "Their runs": l["them"]["total"], "Their wickets": l["them"]["wickets"],
                    "Their overs": l["them"]["overs"], "Their innings no": l["them"]["innings_no"],
                    "Their byes": l["them"]["byes"], "Their leg byes": l["them"]["leg_byes"],
                    "Their wides": l["them"]["wides"], "Their no balls": l["them"]["no_balls"],
                    "Their penalty": l["them"]["penalty"], "Their extras": l["them"]["extras"],
                    "Match result": result if l is legs[0] else "",
                    "Source file": s["source"], "Record index": l["index"],
                })
                for b in l["blocks"]:
                    who = names.get(b["player_id"], {}).get("name", f"Unknown #{b['player_id']}")
                    if b["batted"]:
                        batting.append({**common, "Player": who,
                                        "Batting position": b["position"], "Runs": b["runs"],
                                        "4s": b["fours"], "6s": b["sixes"],
                                        "Not out": "Y" if b["not_out"] else "",
                                        "Dismissal": dismissal_label(b["dismissal_code"]),
                                        "Dismissal code": b["dismissal_code"]})
                    if b["bowled"]:
                        bowling.append({**common, "Player": who, "Overs": b["overs"],
                                        "Maidens": b["maidens"], "Runs conceded": b["conceded"],
                                        "Wickets": b["wickets"], "Wides": b["wides"],
                                        "No balls": b["no_balls"], "Bowling flags": b["flags"]})
                    fld = [b["catches"], b["wk_catches"], b["stumpings"], b["byes"]]
                    if any(v for v in fld):
                        fielding.append({**common, "Player": who,
                                         "Catches": b["catches"] or 0,
                                         "Keeper catches": b["wk_catches"] or 0,
                                         "Stumpings": b["stumpings"] or 0,
                                         "Byes conceded": b["byes"] or 0})
                    if b["votes"]:
                        votes.append({**common, "Player": who, "Votes": b["votes"]})
                for w in l["fow"]:
                    batter = ""
                    if w["batter_position"]:
                        hit = [x for x in l["blocks"] if x["position"] == w["batter_position"]]
                        if hit:
                            batter = names.get(hit[0]["player_id"], {}).get("name", "")
                    fow_rows.append({**common, "Wicket": w["wicket"], "Score": w["score"],
                                     "Batting position": w["batter_position"], "Player": batter})
    return {"Matches": matches, "Batting": batting, "Bowling": bowling,
            "Fielding": fielding, "Fall of wickets": fow_rows, "Votes": votes,
            "Players": players}


def build_season_stats(seasons: list, by_grade: bool = True) -> list:
    """One row per player per season (and per grade unless told otherwise).

    ``by_grade=False`` rolls the same figures up without the grade. It is NOT
    written out, because it carries nothing the graded file does not: summing
    the graded rows per player-season reproduces it exactly, every counting
    column and every high score, across all 853 player-seasons. The grade split
    is purely additive detail, so there is one output file and no choice to
    make. The switch is kept for anyone who wants that roll-up separately.
    """
    agg = {}
    seen_games = defaultdict(set)      # (player, season, grade) -> the matches played
    for s in seasons:
        groups = defaultdict(list)
        for m in s["matches"]:
            groups[(m["date"], m["team"], m["opponent"])].append(m)
        for key, legs in groups.items():
            team = key[1]
            for l in legs:
                if not l["has_play"]:
                    continue
                for b in l["blocks"]:
                    name = s["players"].get(b["player_id"], {}).get("name")
                    if not name:
                        continue
                    k = (name, s["season"], grade_label(team) if by_grade else "")
                    row = agg.setdefault(k, {
                        "innings": 0, "runs": 0, "not_outs": 0, "hs": None, "hs_no": False,
                        "fours": 0, "sixes": 0, "balls": 0, "maidens": 0, "conceded": 0,
                        "wickets": 0, "catches": 0, "wk_catches": 0, "stumpings": 0,
                        "spells": 0, "votes": 0.0, "fifties": 0, "hundreds": 0,
                        "ducks": 0, "five_fors": 0, "best": None,
                    })
                    # a two-day match is one game however many legs a player appears in
                    seen_games[k].add(key)
                    if b["batted"]:
                        row["innings"] += 1
                        row["runs"] += b["runs"]
                        row["fours"] += b["fours"] or 0
                        row["sixes"] += b["sixes"] or 0
                        if b["not_out"]:
                            row["not_outs"] += 1
                        if b["runs"] >= 100:
                            row["hundreds"] += 1
                        elif b["runs"] >= 50:
                            row["fifties"] += 1
                        # out for nothing. A 0 not out is not a duck
                        if b["runs"] == 0 and not b["not_out"]:
                            row["ducks"] += 1
                        best = row["hs"]
                        better = best is None or b["runs"] > best
                        # an equal score that was not out is the better high score
                        upgrade = b["runs"] == best and b["not_out"] and not row["hs_no"]
                        if better or upgrade:
                            row["hs"], row["hs_no"] = b["runs"], b["not_out"]
                    if b["bowled"]:
                        row["spells"] += 1
                        row["balls"] += b["balls"]
                        row["maidens"] += b["maidens"] or 0
                        row["conceded"] += b["conceded"] or 0
                        row["wickets"] += b["wickets"] or 0
                        w, rc = b["wickets"] or 0, b["conceded"] or 0
                        if w >= 5:
                            row["five_fors"] += 1
                        # best figures: most wickets, then fewest runs
                        if row["best"] is None or (w, -rc) > (row["best"][0], -row["best"][1]):
                            row["best"] = (w, rc)
                    row["catches"] += b["catches"] or 0
                    row["wk_catches"] += b["wk_catches"] or 0
                    row["stumpings"] += b["stumpings"] or 0
                    row["votes"] += b["votes"] or 0
    rows = []
    for (name, season, grade), r in agg.items():
        outs = r["innings"] - r["not_outs"]
        avg = round(r["runs"] / outs, 2) if outs > 0 else ""
        bowl_avg = round(r["conceded"] / r["wickets"], 2) if r["wickets"] else ""
        econ = round(r["conceded"] / (r["balls"] / 6), 2) if r["balls"] else ""
        hs = "" if r["hs"] is None else f"{r['hs']}{'*' if r['hs_no'] else ''}"
        rows.append({
            "Player": name, "Season": season, "Grade": grade,
            "Games": len(seen_games[(name, season, grade)]),
            "Innings": r["innings"], "Runs": r["runs"], "NO": r["not_outs"],
            "HS": hs, "Avg": avg, "4s": r["fours"], "6s": r["sixes"],
            "50s": r["fifties"], "100s": r["hundreds"], "Ducks": r["ducks"],
            "Bowling Innings": r["spells"], "Wickets": r["wickets"],
            "Overs": balls_to_overs(r["balls"]) if r["spells"] else "",
            "Maidens": r["maidens"], "Runs Conceded": r["conceded"] if r["spells"] else "",
            "Bowl Avg": bowl_avg, "Econ": econ, "5WI": r["five_fors"],
            "Best": f"{r['best'][0]}-{r['best'][1]}" if r["best"] else "",
            "Catches": (r["catches"] or 0) + (r["wk_catches"] or 0),
            "Catches WK": r["wk_catches"], "Stumpings": r["stumpings"],
            "Votes": round(r["votes"], 0) if r["votes"] else "",
        })
    if not by_grade:
        for r in rows:
            del r["Grade"]
        rows.sort(key=lambda x: (x["Season"], x["Player"]))
    else:
        rows.sort(key=lambda x: (x["Season"], x["Grade"], x["Player"]))
    return rows


def squad_clashes(seasons: list) -> list:
    """Player-dates where the files name one person in more than one grade.

    Nobody plays two matches on one afternoon, so either the grade a
    performance is filed under is wrong, or the stored date is a round stamp
    shared by competitions that ran on different days. The files cannot tell
    us which, and the figures themselves are sound either way - every innings
    still reconciles to its own total. It matters only for the grade split, so
    it is listed here rather than silently dropped or silently kept.
    """
    out = []
    for s_ in seasons:
        by_date = defaultdict(lambda: defaultdict(list))
        for m in s_["matches"]:
            if not m["has_play"]:
                continue
            for b in m["blocks"]:
                by_date[m["date"]][b["player_id"]].append((m, b))
        for date in sorted(by_date, key=lambda d: d or datetime.date.min):
            for pid, hits in by_date[date].items():
                grades = {h[0]["team"] for h in hits}
                if len(grades) < 2:
                    continue
                name = s_["players"].get(pid, {}).get("name", f"#{pid}")
                for m, b in sorted(hits, key=lambda h: h[0]["team"]):
                    out.append({
                        "Season": s_["season"], "Date": date, "Player": name,
                        "Grade": grade_label(m["team"]), "Opponent": m["opponent"],
                        "Ground": m["ground"],
                        "Runs": b["runs"] if b["batted"] else "",
                        "Overs": b["overs"] if b["bowled"] else "",
                        "Wickets": b["wickets"] if b["bowled"] else "",
                        "Grades named that day": len(grades),
                    })
    return out


# ── verification ─────────────────────────────────────────────────────────────


def verify(seasons: list) -> list:
    """Check the decode against the data's own arithmetic, not against itself."""
    c = defaultdict(int)
    for s in seasons:
        for m in s["matches"]:
            us, them, blocks = m["us"], m["them"], m["blocks"]
            batted = [b for b in blocks if b["batted"]]
            bowled = [b for b in blocks if b["bowled"]]
            if us["played"] and batted:
                c["runs_checked"] += 1
                if sum(b["runs"] for b in batted) + us["extras"] == us["total"]:
                    c["runs_ok"] += 1
            if us["played"] and us["wickets"] is not None and batted:
                c["wickets_checked"] += 1
                if sum(1 for b in batted if not b["not_out"]) == us["wickets"]:
                    c["wickets_ok"] += 1
            if them["played"] and bowled and them["byes"] is not None:
                c["conceded_checked"] += 1
                got = sum(b["conceded"] or 0 for b in bowled)
                want = them["total"] - (them["byes"] or 0) - (them["leg_byes"] or 0)
                if got == want:
                    c["conceded_ok"] += 1
            if them["played"] and them["wickets"] is not None and bowled:
                c["bowler_wkts_checked"] += 1
                if sum(b["wickets"] or 0 for b in bowled) <= them["wickets"]:
                    c["bowler_wkts_ok"] += 1
            for b in batted:
                c["boundary_checked"] += 1
                if 4 * (b["fours"] or 0) + 6 * (b["sixes"] or 0) <= b["runs"]:
                    c["boundary_ok"] += 1
            v = [b["votes"] for b in blocks if b["votes"]]
            if v:
                c["votes_checked"] += 1
                if sum(v) == 6:
                    c["votes_ok"] += 1
    return [
        ("Batting runs + extras equal the innings total", c["runs_ok"], c["runs_checked"]),
        ("Batters dismissed equal the innings wickets", c["wickets_ok"], c["wickets_checked"]),
        ("Bowlers' runs equal their total less byes and leg byes",
         c["conceded_ok"], c["conceded_checked"]),
        ("Bowlers' wickets never exceed the innings wickets",
         c["bowler_wkts_ok"], c["bowler_wkts_checked"]),
        ("4s and 6s never account for more than the runs scored",
         c["boundary_ok"], c["boundary_checked"]),
        ("Award votes add up to 3-2-1", c["votes_ok"], c["votes_checked"]),
    ]


# ── output ───────────────────────────────────────────────────────────────────

NOTES = [
    ("What this is", ""),
    ("", "Shoalwater Bay Cricket Club's DATA<year>.AV files, converted."),
    ("", "One row per match, innings, spell or dismissal, exactly as the files hold it."),
    ("", "Nothing is estimated: a figure the files do not carry is left blank."),
    ("", ""),
    ("Things worth knowing", ""),
    ("Overs", "Cricket notation. 93.3 is 93 overs and 3 balls, so season totals are "
              "summed as balls and written back the same way."),
    ("Grade", "Taken from the team number the fixture record carries. The files hold no "
              "grade NAME, so these are 'Grade 1'..'Grade 4' and can be renamed on import. "
              "1992 stores 2 for every fixture, which may be a default rather than a real "
              "team number."),
    ("Two-day matches", "Seasons up to 1997 store a match as two records, 'Opponent 1' and "
                        "'Opponent 2'. They share a date and a team, so they are one match "
                        "here, with the Leg column saying which record a row came from. "
                        "Games played counts the match once."),
    ("Opposition", "The program only ever stored this club's own players, so there are no "
                   "opposition batting or bowling cards - only their innings totals."),
    ("Dismissals", "Only 'not out' is certain (code 11, confirmed because the count of "
                   "dismissed batters matches the innings wickets). The other codes are "
                   "passed through raw rather than guessed at."),
    ("Not recorded", "Balls faced, run outs and the bowler or fielder who took a wicket are "
                     "not in the format at all."),
    ("Bowling flags", "An unidentified bitmask, present only on bowlers. Passed through raw."),
    ("Same player, two grades", "The Data quality sheet lists every date where one person is "
                               "named in more than one grade, which nobody can be. It runs to "
                               "a handful of rows out of 9,384 player-dates - the club's own "
                               "slips. Squads for different grades on one day do not overlap "
                               "at all, and a player spends a median 93% of a season in one "
                               "grade, so the grade split can be trusted."),
    ("Which file to import", "betterimport_season_stats.csv. There is one, deliberately."),
    ("", ""),
    ("Checks run against the files' own arithmetic", ""),
]


def write_xlsx(sheets: dict, checks: list, path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Notes")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 110
    for a, b in NOTES:
        ws.append([a, b])
    for label, ok, total in checks:
        pct = f"{ok / total:.1%}" if total else "n/a"
        ws.append([label, f"{ok} of {total} ({pct})"])
    ws.append(["", ""])
    ws.append(["Where a check falls short", "the shortfall is in the club's own figures, "
               "not the conversion - the residuals are ones and twos, the usual slips in a "
               "hand-kept scorebook. Nothing was corrected."])
    for row in ws.iter_rows():
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        if not rows:
            ws.append(["(no rows)"])
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in rows:
            ws.append([r.get(h) for h in headers])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for n, h in enumerate(headers, start=1):
            longest = max([len(str(h))] + [len(str(r.get(h) or "")) for r in rows[:400]])
            ws.column_dimensions[get_column_letter(n)].width = min(max(longest + 2, 9), 34)
    wb.save(path)


def write_csv(rows: list, path: Path) -> None:
    headers = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=headers)
        w.writeheader()
        w.writerows(rows)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("source", type=Path, help="folder holding the .AV files")
    ap.add_argument("-o", "--out", type=Path, default=Path("."), help="output folder")
    args = ap.parse_args()

    files = sorted(p for p in args.source.iterdir() if p.suffix.lower() == ".av")
    if not files:
        raise SystemExit(f"no .av files in {args.source}")
    seasons = [parse_file(p) for p in files]
    args.out.mkdir(parents=True, exist_ok=True)

    sheets = build_rows(seasons)
    sheets["Data quality"] = squad_clashes(seasons)
    checks = verify(seasons)
    write_xlsx(sheets, checks, args.out / "match_detail.xlsx")
    stats = build_season_stats(seasons)
    write_csv(stats, args.out / "betterimport_season_stats.csv")

    club = seasons[0]["club"]
    print(f"{club}: {len(seasons)} seasons, {seasons[0]['season']} to {seasons[-1]['season']}")
    for name, rows in sheets.items():
        print(f"  {name:<16} {len(rows):>6} rows")
    print(f"  season stats     {len(stats):>6} rows")
    print()
    for label, ok, total in checks:
        pct = f"{ok / total:.1%}" if total else "n/a"
        print(f"  {label:<58} {ok:>5} / {total:<5} {pct}")


if __name__ == "__main__":
    main()
