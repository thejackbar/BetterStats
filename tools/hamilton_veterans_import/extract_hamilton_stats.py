"""Turn Hamilton Veterans' season-statistics workbook into BetterImport CSVs.

The club keeps its history in one workbook whose history tabs are block
structured, not tabular: a player's name sits alone in column A, their seasons
run down column B underneath it, and the stat columns repeat once per
competition (VCV, Peter Phyland Trophy, Border Cup, One Off Games, WillowFest,
Echuca Carnival) under a merged banner row. A "Total" row closes each block.

BetterImport (`/admin/import`) wants one flat row per player per season, so this
reads the blocks, joins batting to bowling to games played, and writes:

  hamilton_import_season_by_season.csv   one row per player/season/competition
  hamilton_import_season_totals.csv      one row per player/season (all comps)
  hamilton_players.csv                   the roster the CSVs name, with debut
  hamilton_seasons.csv                   the seasons to create before importing
  hamilton_data_notes.csv                every figure worth a second look

Guest players (the "Extras" tabs, where a name carries their own club in
brackets) are written to their own `_guests` copies rather than mixed in — they
turned out for Hamilton but they are somebody else's registered player, and a
club deciding whether to carry them in its own records is a judgement call.

Run:  python3 tools/hamilton_veterans_import/extract_hamilton_stats.py <xlsx> [outdir]
"""

from __future__ import annotations

import csv
import re
import sys
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl

# ── sheet layout ─────────────────────────────────────────────────────────────

# Every competition group repeats the same columns, so a group is described by
# its first column and the fields in order. Column numbers are 1-based.
BATTING_COLS = ["inns", "retired_no", "not_out", "out", "runs", "balls", "aver", "sr", "fours", "sixes", "hs"]
BOWLING_COLS = ["overs", "maidens", "wickets", "runs", "aver", "sr", "econ", "wides", "no_balls", "best"]

BATTING_SHEETS = [
    # (sheet, first data row, first group column, group width, competition names, guests?)
    ("BattingHistory A-M", 4, 3, 11, None, False),
    ("BattingHistory N-Z", 4, 3, 11, None, False),
    ("BattingHistory Extras", 4, 10, 11, None, True),
]
BOWLING_SHEETS = [
    ("Bowling History A-M", 4, 3, 10, None, False),
    ("BowlingHistory N-Z", 4, 3, 10, None, False),
    ("BowlingHistory Extras", 4, 3, 10, None, True),
]

# The banner row spells the same competition several ways across tabs.
COMP_CANONICAL = {
    "vcv": "VCV",
    "vosca / vcv": "VCV",
    "peter phyland trophy": "Peter Phyland Trophy",
    "border cup": "Border Cup",
    "one off games": "One Off Games",
    "one off 40 over games": "One Off Games",
    "one off 20 over games": "One Off Games",
    "nat o60s champs": "One Off Games",
    "national o60s champs": "One Off Games",
    "portland tourn": "One Off Games",
    "mildura willowfest": "WillowFest",
    "willowfest - riverland / mildura": "WillowFest",
    "willow fest": "WillowFest",
    "willow fest (rivrland / mildura)": "WillowFest",
    "echuca carnival": "Echuca Carnival",
}
COMP_ORDER = ["VCV", "Peter Phyland Trophy", "Border Cup", "One Off Games", "WillowFest", "Echuca Carnival"]

GAMES_SHEET = "Games Played Total"
GAMES_DETAIL_FIRST_COL = 3      # VOSCA/VCV … Echuca Carnival, then Total
GAMES_DETAIL_LAST_COL = 11
GAMES_SUMMARY_NAME_COL = 16
GAMES_SUMMARY_DEBUT_COL = 17

TOTAL_LABELS = {"total", "totals", "career", "career total"}
JUNK = {"", "#div/0!", "#value!", "#ref!", "#n/a", "-", "none"}


# ── cell helpers ─────────────────────────────────────────────────────────────

def txt(v) -> str:
    return "" if v is None else str(v).strip()


def num(v):
    """A cell as a float, or None for blanks and Excel error strings."""
    s = txt(v)
    if s.lower() in JUNK:
        return None
    s = s.replace("*", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def whole(v):
    f = num(v)
    return None if f is None else int(round(f))


def overs_to_balls(overs) -> int:
    """Cricket notation in, balls out: 26.2 is 26 overs and 2 balls = 158."""
    if overs is None:
        return 0
    o = int(overs)
    balls = int(round((float(overs) - o) * 10))
    if balls > 5:                      # 26.7 is not a legal figure; treat as overs
        o, balls = o + balls // 6, balls % 6
    return o * 6 + balls


def balls_to_overs(balls: int) -> str:
    o, b = divmod(int(balls), 6)
    return f"{o}.{b}" if b else str(o)


def best_figures(v) -> str:
    """'3/9' → '3-9'. Anything that is not W/R is dropped rather than guessed."""
    s = txt(v).replace("*", "").strip()
    m = re.match(r"^(\d+)\s*[/-]\s*(\d+)$", s)
    return f"{m.group(1)}-{m.group(2)}" if m else ""


def best_rank(fig: str):
    """Sort key for 'W-R' — most wickets, then fewest runs."""
    m = re.match(r"^(\d+)-(\d+)$", fig or "")
    return (int(m.group(1)), -int(m.group(2))) if m else (-1, 0)


# ── names and seasons ────────────────────────────────────────────────────────

GUEST_RE = re.compile(r"\s*\[([^\]]*)\]?\s*$")     # the closing bracket is sometimes missing
INITIAL_RE = re.compile(r"^[A-Z]$|^[A-Z]\.$")


def split_guest(raw: str):
    """'Baulch, Colin [Geelong]' → ('Baulch, Colin', 'Geelong')."""
    m = GUEST_RE.search(raw)
    if not m:
        return raw.strip(), ""
    return GUEST_RE.sub("", raw).strip(), m.group(1).strip()


def display_name(raw: str) -> str:
    """The sheet writes 'Surname, First' (mostly) — we want 'First Surname'.

    It also writes 'Skrabl. Peter' with a full stop for the comma, 'Barker
    Geoff' with nothing at all, and 'Kearney, ???' where the first name was
    never known.
    """
    s = re.sub(r"\s+", " ", raw.replace("?", " ")).strip(" ,.")
    s = re.sub(r"^([A-Za-z'\-]{2,})\.\s+", r"\1, ", s)      # 'Skrabl. Peter'
    if "," in s:
        last, first = [p.strip(" .") for p in s.split(",", 1)]
        return f"{first} {last}".strip() if first else last
    parts = s.split(" ")
    if len(parts) == 2 and not INITIAL_RE.match(parts[1]):
        # 'Barker Geoff' — the comma was forgotten, surname still comes first.
        return f"{parts[1]} {parts[0]}"
    return s


def name_key(raw: str) -> str:
    """Match one person across tabs however their name was typed."""
    s = split_guest(raw)[0].lower()
    s = re.sub(r"[^a-z ]+", " ", s)
    return " ".join(sorted(t for t in s.split() if t))


def season_label(raw: str) -> str:
    """'18/19' → 'Summer 2018/19', matching how a season is named in the app.

    Returns "" for anything that is not one summer — including '13/15', which
    the workbook carries in a couple of places and which spans two.
    """
    s = txt(raw)
    m = re.match(r"^(\d{2})\s*/\s*(\d{2})$", s)
    if not m:
        return ""
    yy, yy2 = int(m.group(1)), int(m.group(2))
    if (yy + 1) % 100 != yy2:
        return ""
    start = yy + (2000 if yy < 50 else 1900)
    return f"Summer {start}/{m.group(2)}"


def season_sort(label: str) -> int:
    m = re.search(r"(\d{4})", label or "")
    return int(m.group(1)) if m else 0


# ── block readers ────────────────────────────────────────────────────────────

def read_groups(ws, first_col: int, width: int, count_hint=None):
    """Competition name per group, read off the merged banner row (row 2)."""
    banner = [txt(c) for c in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
    groups = []
    col = first_col
    while col <= ws.max_column - width + 1:
        raw = banner[col - 1] if col - 1 < len(banner) else ""
        key = re.sub(r"\s+", " ", raw.strip().lower())
        comp = COMP_CANONICAL.get(key)
        if comp is None:
            # An unlabelled trailing group is the sheet's own Total column set.
            groups.append((col, "__total__"))
        else:
            groups.append((col, comp))
        col += width
    return groups


def read_blocks(ws, first_row: int, first_col: int, width: int, notes=None, sheet=""):
    """Yield (raw_player_name, [(season_cell, row_values)], total_row_values).

    A block ends at its FIRST 'Total' row. Every history tab carries spare,
    unnamed template blocks below the last real player — season rows and their
    own zeroed Total row, with nobody's name on them. Reading past the Total
    row would hang those on whoever was named last, and their zeroed Total
    would then overwrite the real one. Anything left over that actually holds a
    figure is reported rather than dropped.
    """
    name = None
    rows: list = []
    total = None
    closed = False

    def stray(row) -> bool:
        """A real figure, not the zeros and #DIV/0! a spare template row holds."""
        for c in range(first_col, min(len(row), first_col + width * 7) + 1):
            if c - 1 >= len(row):
                break
            v = num(row[c - 1])
            if v:
                return True
            if v is None and best_figures(row[c - 1]):
                return True
        return False

    for row in ws.iter_rows(min_row=first_row, values_only=True):
        who, season = txt(row[0]), txt(row[1])
        if who:
            if name:
                yield name, rows, total
            name, rows, total, closed = who, [], None, False
            if season and season.lower() not in TOTAL_LABELS:
                rows.append((season, row))          # a name and a season on the one row
            continue
        if not name:
            continue
        if closed:
            if season and stray(row) and notes is not None:
                notes.append([sheet, name, season,
                              "figures sit in an unnamed block below a player's total row — nobody to credit them to"])
            continue
        if season.lower() in TOTAL_LABELS:
            total = row
            closed = True
        elif season:
            rows.append((season, row))
    if name:
        yield name, rows, total


def group_values(row, col: int, fields: list) -> dict:
    out = {}
    for i, f in enumerate(fields):
        idx = col - 1 + i
        out[f] = row[idx] if idx < len(row) else None
    return out


# ── parsing ──────────────────────────────────────────────────────────────────

def parse_batting(wb, notes):
    """{(key, season, comp): stats}, plus the sheet's own per-player totals."""
    data: dict = {}
    sheet_totals: dict = {}
    names: dict = {}
    for sheet, first_row, first_col, width, _c, guest in BATTING_SHEETS:
        ws = wb[sheet]
        groups = read_groups(ws, first_col, width)
        for raw, rows, total in read_blocks(ws, first_row, first_col, width, notes, sheet):
            key = name_key(raw)
            names.setdefault(key, []).append((raw, sheet, guest))
            for season_cell, row in rows:
                label = season_label(season_cell)
                if not label:
                    notes.append([sheet, raw, txt(season_cell), "unreadable season label — row skipped"])
                    continue
                for col, comp in groups:
                    if comp == "__total__":
                        continue
                    g = group_values(row, col, BATTING_COLS)
                    inns = whole(g["inns"])
                    runs = whole(g["runs"])
                    if not inns and not runs:
                        continue
                    rt_no, no, out = whole(g["retired_no"]) or 0, whole(g["not_out"]) or 0, whole(g["out"]) or 0
                    inns = inns or 0
                    if inns and rt_no + no + out and inns != rt_no + no + out:
                        notes.append([sheet, raw, f"{label} / {comp}",
                                      f"innings {inns} does not equal retired-not-out {rt_no} + not out {no} + out {out}"])
                    rec = data.setdefault((key, label, comp), _blank_bat())
                    rec["batting_innings"] += inns
                    rec["batting_not_outs"] += rt_no + no
                    rec["batting_outs"] += out
                    rec["batting_runs"] += runs or 0
                    rec["batting_balls"] += whole(g["balls"]) or 0
                    rec["batting_fours"] += whole(g["fours"]) or 0
                    rec["batting_sixes"] += whole(g["sixes"]) or 0
                    hs = whole(g["hs"])
                    if hs is not None and hs > (rec["batting_high_score"] if rec["batting_high_score"] is not None else -1):
                        rec["batting_high_score"] = hs
                        rec["hs_all_not_out"] = inns > 0 and out == 0
            if total is not None:
                acc = _blank_bat()
                for col, comp in groups:
                    if comp == "__total__":
                        continue
                    g = group_values(total, col, BATTING_COLS)
                    acc["batting_innings"] += whole(g["inns"]) or 0
                    acc["batting_runs"] += whole(g["runs"]) or 0
                sheet_totals.setdefault(key, {"batting_innings": 0, "batting_runs": 0})
                sheet_totals[key]["batting_innings"] += acc["batting_innings"]
                sheet_totals[key]["batting_runs"] += acc["batting_runs"]
    return data, sheet_totals, names


def _blank_bat():
    return {
        "batting_innings": 0, "batting_not_outs": 0, "batting_outs": 0, "batting_runs": 0,
        "batting_balls": 0, "batting_fours": 0, "batting_sixes": 0,
        "batting_high_score": None, "hs_all_not_out": False,
    }


def _blank_bowl():
    return {
        "bowling_balls": 0, "bowling_maidens": 0, "bowling_wickets": 0, "bowling_runs": 0,
        "bowling_wides": 0, "bowling_no_balls": 0, "bowling_best_figures": "",
    }


def parse_bowling(wb, notes):
    data: dict = {}
    sheet_totals: dict = {}
    names: dict = {}
    for sheet, first_row, first_col, width, _c, guest in BOWLING_SHEETS:
        ws = wb[sheet]
        groups = read_groups(ws, first_col, width)
        for raw, rows, total in read_blocks(ws, first_row, first_col, width, notes, sheet):
            key = name_key(raw)
            names.setdefault(key, []).append((raw, sheet, guest))
            for season_cell, row in rows:
                label = season_label(season_cell)
                if not label:
                    notes.append([sheet, raw, txt(season_cell), "unreadable season label — row skipped"])
                    continue
                for col, comp in groups:
                    if comp == "__total__":
                        continue
                    g = group_values(row, col, BOWLING_COLS)
                    overs, wkts, runs = num(g["overs"]), whole(g["wickets"]), whole(g["runs"])
                    if not overs and not wkts and not runs:
                        continue
                    balls = overs_to_balls(overs)
                    rec = data.setdefault((key, label, comp), _blank_bowl())
                    rec["bowling_balls"] += balls
                    rec["bowling_maidens"] += whole(g["maidens"]) or 0
                    rec["bowling_wickets"] += wkts or 0
                    rec["bowling_runs"] += runs or 0
                    rec["bowling_wides"] += whole(g["wides"]) or 0
                    rec["bowling_no_balls"] += whole(g["no_balls"]) or 0
                    fig = best_figures(g["best"])
                    if fig and best_rank(fig) > best_rank(rec["bowling_best_figures"]):
                        rec["bowling_best_figures"] = fig
                    if wkts and fig and int(fig.split("-")[0]) > wkts:
                        notes.append([sheet, raw, f"{label} / {comp}",
                                      f"best figures {fig} claim more wickets than the {wkts} recorded for the season"])
            if total is not None:
                acc = {"bowling_wickets": 0, "bowling_runs": 0}
                for col, comp in groups:
                    if comp == "__total__":
                        continue
                    g = group_values(total, col, BOWLING_COLS)
                    acc["bowling_wickets"] += whole(g["wickets"]) or 0
                    acc["bowling_runs"] += whole(g["runs"]) or 0
                st = sheet_totals.setdefault(key, {"bowling_wickets": 0, "bowling_runs": 0})
                st["bowling_wickets"] += acc["bowling_wickets"]
                st["bowling_runs"] += acc["bowling_runs"]
    return data, sheet_totals, names


def parse_games(wb, notes):
    """{(key, season, comp): games} from the detailed block, plus debut dates."""
    ws = wb[GAMES_SHEET]
    header = [txt(c) for c in next(ws.iter_rows(min_row=3, max_row=3, values_only=True))]
    comps = {}
    for col in range(GAMES_DETAIL_FIRST_COL, GAMES_DETAIL_LAST_COL + 1):
        key = re.sub(r"\s+", " ", header[col - 1].strip().lower())
        comp = COMP_CANONICAL.get(key)
        if comp:
            comps[col] = comp
        else:
            notes.append([GAMES_SHEET, "", header[col - 1], "games-played column matched no competition — ignored"])

    games: dict = defaultdict(int)
    seen_names: dict = {}
    for raw, rows, _total in read_blocks(ws, 4, GAMES_DETAIL_FIRST_COL, GAMES_DETAIL_LAST_COL - GAMES_DETAIL_FIRST_COL + 1, notes, GAMES_SHEET):
        key = name_key(raw)
        seen_names.setdefault(key, raw)
        for season_cell, row in rows:
            label = season_label(season_cell)
            if not label:
                continue
            for col, comp in comps.items():
                n = whole(row[col - 1]) if col - 1 < len(row) else None
                if n:
                    games[(key, label, comp)] += n

    debut = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        raw = txt(row[GAMES_SUMMARY_NAME_COL - 1]) if GAMES_SUMMARY_NAME_COL - 1 < len(row) else ""
        if not raw:
            continue
        d = row[GAMES_SUMMARY_DEBUT_COL - 1] if GAMES_SUMMARY_DEBUT_COL - 1 < len(row) else None
        val = ""
        if hasattr(d, "strftime"):
            val = d.strftime("%Y-%m-%d")
        elif txt(d) and txt(d).lower() != "unknown date":
            val = txt(d)
        debut[name_key(raw)] = val
        seen_names.setdefault(name_key(raw), raw)
    return dict(games), debut, seen_names


def align_game_names(games: dict, known: set, spellings: dict, notes: list) -> dict:
    """Fix a name the Games Played tab spells differently from the history tabs.

    'Asbhy, Gary' and 'Ashby, Gary' are one person with a typo between them. The
    test is deliberately narrow: same number of name parts, all but one part
    identical, and the odd pair at least three letters long and close. That last
    condition is what keeps 'Wilson, R' and 'Wilson, T' apart — two people whose
    names differ by a single initial are exactly the pair a looser rule merges.
    """
    remap: dict = {}
    for key in {k for k, _l, _c in games}:
        if key in known:
            continue
        parts = key.split()
        best = None
        for cand in known:
            cp = cand.split()
            if len(cp) != len(parts):
                continue
            diff = [(a, b) for a, b in zip(sorted(parts), sorted(cp)) if a != b]
            if len(diff) != 1:
                continue
            a, b = diff[0]
            if min(len(a), len(b)) < 3:
                continue
            score = SequenceMatcher(None, a, b).ratio()
            if score >= 0.8 and (best is None or score > best[1]):
                best = (cand, score)
        if best:
            remap[key] = best[0]
            notes.append(["names", display_name(sorted(spellings.get(best[0], [best[0]]))[0]),
                          f"{key} → {best[0]}",
                          "the Games Played tab spells this player's name differently — read as the same person"])
    if not remap:
        return games
    out: dict = defaultdict(int)
    for (key, label, comp), n in games.items():
        out[(remap.get(key, key), label, comp)] += n
    return dict(out)


# ── assembly ─────────────────────────────────────────────────────────────────

ROW_FIELDS = [
    "player_name", "season_label", "grade_label", "games_played",
    "batting_innings", "batting_not_outs", "batting_runs", "batting_balls",
    "batting_fours", "batting_sixes", "batting_high_score",
    "bowling_overs", "bowling_maidens", "bowling_runs", "bowling_wickets",
    "bowling_wides", "bowling_no_balls", "bowling_best_figures",
]


def build_rows(batting, bowling, games, display, guests):
    keys = set(batting) | set(bowling) | set(games)
    out = []
    for key, label, comp in sorted(keys, key=lambda k: (display.get(k[0], k[0]), season_sort(k[1]), COMP_ORDER.index(k[2]) if k[2] in COMP_ORDER else 99)):
        bat = batting.get((key, label, comp), _blank_bat())
        bowl = bowling.get((key, label, comp), _blank_bowl())
        g = games.get((key, label, comp), 0)
        hs = bat["batting_high_score"]
        row = {
            "player_name": display.get(key, key),
            "season_label": label,
            "grade_label": comp,
            "games_played": g or "",
            "batting_innings": bat["batting_innings"] or "",
            "batting_not_outs": bat["batting_not_outs"] or "",
            "batting_runs": bat["batting_runs"] or "",
            "batting_balls": bat["batting_balls"] or "",
            "batting_fours": bat["batting_fours"] or "",
            "batting_sixes": bat["batting_sixes"] or "",
            "batting_high_score": ("" if hs is None else f"{hs}*" if bat["hs_all_not_out"] and hs else hs),
            "bowling_overs": balls_to_overs(bowl["bowling_balls"]) if bowl["bowling_balls"] else "",
            "bowling_maidens": bowl["bowling_maidens"] or "",
            "bowling_runs": bowl["bowling_runs"] or "",
            "bowling_wickets": bowl["bowling_wickets"] or "",
            "bowling_wides": bowl["bowling_wides"] or "",
            "bowling_no_balls": bowl["bowling_no_balls"] or "",
            "bowling_best_figures": bowl["bowling_best_figures"],
        }
        if any(row[f] not in ("", 0) for f in ROW_FIELDS[3:]):
            out.append((key, row))
    return out


SUMMABLE = (
    "games_played", "batting_innings", "batting_not_outs", "batting_runs", "batting_balls",
    "batting_fours", "batting_sixes", "bowling_maidens", "bowling_runs", "bowling_wickets",
    "bowling_wides", "bowling_no_balls",
)


def collapse_to_season(rows):
    """One row per player and season, competitions summed.

    Overs are summed as BALLS and converted back, since 26.2 + 13.5 is 40.1 in
    cricket notation, not 39.7.
    """
    merged: dict = {}
    order: list = []
    for key, r in rows:
        k = (key, r["season_label"])
        if k not in merged:
            merged[k] = {
                "player_name": r["player_name"], "season_label": r["season_label"], "grade_label": "",
                "batting_high_score": "", "bowling_best_figures": "", "_balls": 0,
                **{f: 0 for f in SUMMABLE},
            }
            order.append(k)
        m = merged[k]
        for f in SUMMABLE:
            m[f] += int(r[f]) if str(r[f]).strip() else 0
        if r["bowling_overs"]:
            m["_balls"] += overs_to_balls(num(r["bowling_overs"]))
        hs_new, hs_old = str(r["batting_high_score"]), str(m["batting_high_score"])
        if hs_new and (not hs_old or int(hs_new.rstrip("*")) > int(hs_old.rstrip("*"))):
            m["batting_high_score"] = r["batting_high_score"]
        if r["bowling_best_figures"] and best_rank(r["bowling_best_figures"]) > best_rank(m["bowling_best_figures"]):
            m["bowling_best_figures"] = r["bowling_best_figures"]

    out = []
    for k in order:
        m = merged[k]
        balls = m.pop("_balls")
        m["bowling_overs"] = balls_to_overs(balls) if balls else ""
        out.append((k[0], {f: ("" if m[f] == 0 else m[f]) for f in ROW_FIELDS}))
    return out


def write_csv(path: Path, fields, rows):
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def main(src: str, outdir: str):
    out = Path(outdir)
    out.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    notes: list = []

    batting, bat_totals, bat_names = parse_batting(wb, notes)
    bowling, bowl_totals, bowl_names = parse_bowling(wb, notes)
    games, debut, game_names = parse_games(wb, notes)

    # One display name per person, preferring a spelling that carries a full
    # first name over one that only has an initial.
    spellings: dict = defaultdict(list)
    guests: dict = {}
    for src_names in (bat_names, bowl_names):
        for key, entries in src_names.items():
            for raw, sheet, is_guest in entries:
                base, club = split_guest(raw)
                spellings[key].append(base)
                if club:
                    guests[key] = club
                elif is_guest:
                    guests.setdefault(key, "")
    display = {}
    for key, opts in spellings.items():
        best = sorted(set(opts), key=lambda s: (-len(re.sub(r"[^A-Za-z]", "", s)), s))[0]
        display[key] = display_name(best)
        if len(set(display_name(o) for o in opts)) > 1:
            notes.append(["names", display[key], " / ".join(sorted(set(opts))),
                          "one player is spelled several ways across tabs — merged under the fullest spelling"])

    games = align_game_names(games, set(spellings), spellings, notes)

    # Games played for someone with no batting or bowling record at all still
    # names a real person; keep them so the roster and the games count agree.
    for key in {k for k, _l, _c in games}:
        raw = game_names.get(key, key.title())
        display.setdefault(key, display_name(raw))
        spellings.setdefault(key, []).append(split_guest(raw)[0])

    rows = build_rows(batting, bowling, games, display, guests)
    season_rows = collapse_to_season(rows)

    def split(pairs):
        club = [r for k, r in pairs if k not in guests]
        guest = [r for k, r in pairs if k in guests]
        return club, guest

    club_rows, guest_rows = split(rows)
    club_seasons, guest_seasons = split(season_rows)

    write_csv(out / "hamilton_import_season_by_season.csv", ROW_FIELDS, club_rows)
    write_csv(out / "hamilton_import_season_totals.csv", ROW_FIELDS, club_seasons)
    write_csv(out / "hamilton_import_season_by_season_guests.csv", ROW_FIELDS, guest_rows)
    write_csv(out / "hamilton_import_season_totals_guests.csv", ROW_FIELDS, guest_seasons)

    # Roster
    player_fields = ["player_name", "sheet_name", "guest_club", "debut", "seasons", "games", "runs", "wickets"]
    per_player: dict = defaultdict(lambda: {"seasons": set(), "games": 0, "runs": 0, "wickets": 0})
    for key, r in rows:
        p = per_player[key]
        p["seasons"].add(r["season_label"])
        for f, t in (("games_played", "games"), ("batting_runs", "runs"), ("bowling_wickets", "wickets")):
            p[t] += int(r[f]) if str(r[f]).strip() else 0
    roster = []
    for key in sorted(per_player, key=lambda k: display.get(k, k)):
        p = per_player[key]
        name = display.get(key, key)
        if len(name.split()) < 2:
            notes.append(["names", name, sorted(set(spellings.get(key, [name])))[0],
                          "no first name on file — worth naming before the import so they are not "
                          "mistaken for another player of the same surname"])
        roster.append({
            "player_name": display.get(key, key),
            "sheet_name": sorted(set(spellings.get(key, [])))[0] if spellings.get(key) else "",
            "guest_club": guests.get(key, ""),
            "debut": debut.get(key, ""),
            "seasons": len(p["seasons"]),
            "games": p["games"], "runs": p["runs"], "wickets": p["wickets"],
        })
    write_csv(out / "hamilton_players.csv", player_fields, roster)

    seasons = sorted({r["season_label"] for _k, r in rows}, key=season_sort)
    write_csv(out / "hamilton_seasons.csv", ["season_name", "year"],
              [{"season_name": s, "year": season_sort(s)} for s in seasons])

    # Cross-check our per-player sums against the workbook's own Total rows.
    for key, tot in bat_totals.items():
        got_i = sum(int(r["batting_innings"] or 0) for k, r in rows if k == key)
        got_r = sum(int(r["batting_runs"] or 0) for k, r in rows if k == key)
        if got_i != tot["batting_innings"] or got_r != tot["batting_runs"]:
            notes.append(["check", display.get(key, key), "batting",
                          f"seasons add to {got_i} innings / {got_r} runs, the workbook's own total row says "
                          f"{tot['batting_innings']} / {tot['batting_runs']}"])
    for key, tot in bowl_totals.items():
        got_w = sum(int(r["bowling_wickets"] or 0) for k, r in rows if k == key)
        got_r = sum(int(r["bowling_runs"] or 0) for k, r in rows if k == key)
        if got_w != tot["bowling_wickets"] or got_r != tot["bowling_runs"]:
            notes.append(["check", display.get(key, key), "bowling",
                          f"seasons add to {got_w} wickets / {got_r} runs, the workbook's own total row says "
                          f"{tot['bowling_wickets']} / {tot['bowling_runs']}"])
    for key, r in rows:
        g = int(r["games_played"] or 0)
        i = int(r["batting_innings"] or 0)
        if g and i > g:
            notes.append(["check", r["player_name"], f"{r['season_label']} / {r['grade_label']}",
                          f"{i} innings against {g} games played"])
        if not g and (i or int(r["bowling_wickets"] or 0)):
            notes.append(["check", r["player_name"], f"{r['season_label']} / {r['grade_label']}",
                          "has batting or bowling but no games recorded on the Games Played tab"])

    write_csv(out / "hamilton_data_notes.csv", ["source", "player", "where", "note"],
              [{"source": n[0], "player": n[1], "where": n[2], "note": n[3]} for n in notes])

    print(f"players            {len(roster)} ({sum(1 for r in roster if not r['guest_club'])} club, "
          f"{sum(1 for r in roster if r['guest_club'])} guest)")
    print(f"seasons            {len(seasons)}  {seasons[0]} … {seasons[-1]}")
    print(f"rows per comp      {len(club_rows)} club, {len(guest_rows)} guest")
    print(f"rows per season    {len(club_seasons)} club, {len(guest_seasons)} guest")
    print(f"runs / wickets     {sum(r['runs'] for r in roster)} / {sum(r['wickets'] for r in roster)}")
    print(f"notes              {len(notes)}")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else "tools/hamilton_veterans_import/out")
