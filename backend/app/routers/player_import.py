"""BetterImport (player profiles) — bulk-fill player contact + profile fields
from a club spreadsheet.

The Players screen lets you edit one player at a time; this fills the same
fields in bulk from a CSV/XLSX: **email** and **phone** (the common case), plus
optional **squad** (selection pool), **role**, **batting hand**, **bowling** and
**gender**. Any column can be left out — a blank cell never overwrites a value a
player already has.

Flow (stateless until commit, same shape as the stats importer):
  1. POST /preview  — upload a file → parsed grid + auto column mapping.
  2. POST /resolve  — {rows, mapping, overrides} → name matches, squad matches,
                      and a per-player "what will change" preview. Idempotent;
                      called again whenever the mapping or a match is adjusted.
  3. POST /commit   — write the changes (and create any new players / squads).

Names are matched with ``import_ingest.match_players`` — the one matcher used
across the site — so exact names link straight through, close ones are offered
as suggestions, and unmatched ones can be created new or pointed at a player.

Gated by MANAGE_PLAYERS, the same capability the Players screen and the
single-player profile editor use.
"""

from __future__ import annotations

import csv
import io
import uuid
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth.capabilities import MANAGE_PLAYERS, require_cap
from app.models.db import Organisation, Player, Team, User, get_db
from app.routers.auth import get_current_club, get_current_user
from app.services import import_ingest as ingest
from app.services import profile_import as prof
from app.services.audit_log import log_activity
from app.services.squad_membership import sync_squad_membership

router = APIRouter(prefix="/club-admin/player-import", tags=["player-import"])

MAX_UPLOAD_BYTES = 8 * 1024 * 1024


# ── request bodies ────────────────────────────────────────────────────────────

class ResolveRequest(BaseModel):
    rows: list[dict] = []
    mapping: dict = {}                  # field -> column header (or {"column": ...})
    # Only used when the name arrives as one combined column (mapping.player_name) —
    # ignored once mapping.player_first_name/player_last_name are mapped instead.
    name_format: str = "auto"           # auto | first_last | last_first
    player_overrides: dict = {}         # raw_name  -> player_id | "__new__" | "__skip__"
    squad_overrides: dict = {}          # raw_squad -> team_id   | "__new__" | "__skip__"


class CommitRequest(ResolveRequest):
    filename: Optional[str] = None


# ── helpers ───────────────────────────────────────────────────────────────────

def _col(mapping: dict, field: str) -> Optional[str]:
    v = mapping.get(field)
    if isinstance(v, dict):
        return v.get("column")
    return v or None


def _row_values(row: dict, mapping: dict) -> dict:
    """{value_field: raw_cell} for every mapped value field present in this row."""
    out = {}
    for field in prof.VALUE_FIELDS:
        col = _col(mapping, field)
        if col is not None and col in row:
            out[field] = row[col]
    return out


def _apply_player_overrides(matches: dict, overrides: dict) -> None:
    for name, choice in (overrides or {}).items():
        if name not in matches:
            matches[name] = {"candidates": []}
        if choice == "__skip__":
            matches[name].update(player_id=None, status="skip")
        elif choice == "__new__":
            matches[name].update(player_id=None, status="new")
        elif choice:
            matches[name].update(player_id=str(choice), status="manual", confidence=1.0)


def _apply_squad_overrides(matches: dict, overrides: dict) -> None:
    for label, choice in (overrides or {}).items():
        if label not in matches:
            matches[label] = {"candidates": []}
        if choice == "__skip__":
            matches[label].update(team_id=None, status="skip")
        elif choice == "__new__":
            matches[label].update(team_id=None, status="new")
        elif choice:
            matches[label].update(team_id=str(choice), status="manual", confidence=1.0)


def _current_profile(p: Player) -> dict:
    """What this player holds today, for the "from → to" preview. Every field
    the importer can write has to be in here, or a sheet re-stating a value a
    player already has reads as a change and inflates the count."""
    return {
        "email": p.email,
        "phone": p.phone,
        "gender": p.gender,
        "player_role": p.player_role,
        "batting_hand": p.batting_hand,
        "bowling_action": p.bowling_action,
        "bowling_type": p.bowling_type,
        "date_of_birth": p.date_of_birth,
        "is_opening_batsman": p.is_opening_batsman,
        "is_overseas": p.is_overseas,
        "overseas_country": p.overseas_country,
        "status": p.status,
        "is_public": p.is_public is not False,
        "is_financial_override": p.is_financial_override,
        "trained_override": p.trained_override,
        "squad_team_id": str(p.squad_team_id) if p.squad_team_id else None,
    }


async def _org_players(db: AsyncSession, org_id) -> list:
    return (await db.execute(select(Player).where(Player.organisation_id == org_id))).scalars().all()


async def _org_teams(db: AsyncSession, org_id) -> list:
    return (await db.execute(select(Team).where(Team.organisation_id == org_id))).scalars().all()


def _squad_choice(label: str, smatch: dict) -> tuple:
    """For a row's raw squad label → ('team', team_id) | ('new', label) |
    ('skip', None) | ('unresolved', label) | ('none', None)."""
    if not label or not str(label).strip():
        return ("none", None)
    m = smatch.get(label) or {}
    st = m.get("status")
    if st == "skip":
        return ("skip", None)
    if st == "new":
        return ("new", label)
    if m.get("team_id"):
        return ("team", m["team_id"])
    return ("unresolved", label)


async def _resolve(db: AsyncSession, org_id, req: ResolveRequest) -> dict:
    """Shared core of /resolve and /commit: match names + squads, then build the
    per-player change preview. Also returns an internal ``_targets`` list the
    commit step writes from (stripped before /resolve returns it)."""
    name_col = _col(req.mapping, "player_name")
    first_col = _col(req.mapping, "player_first_name")
    last_col = _col(req.mapping, "player_last_name")
    if not name_col and not (first_col or last_col):
        raise HTTPException(422, "No column is mapped to the player name.")
    squad_col = _col(req.mapping, "squad")

    def row_name(r):
        return ingest.resolve_row_name(r, name_col=name_col, first_col=first_col,
                                        last_col=last_col, name_format=req.name_format)

    # Distinct names + squad labels, in first-seen order.
    names: list = []
    squad_labels: list = []
    for r in req.rows:
        nm = row_name(r)
        if nm and nm not in names:
            names.append(nm)
        if squad_col:
            sq = str(r.get(squad_col, "")).strip()
            if sq and sq not in squad_labels:
                squad_labels.append(sq)

    players = await _org_players(db, org_id)
    player_by_id = {str(p.id): p for p in players}
    pmatch = ingest.match_players(names, [(str(p.id), p.display_name) for p in players])
    _apply_player_overrides(pmatch, req.player_overrides)

    teams = await _org_teams(db, org_id)
    team_name_by_id = {str(t.id): t.name for t in teams}
    smatch = prof.match_squads(squad_labels, [(str(t.id), t.name) for t in teams]) if squad_col else {}
    _apply_squad_overrides(smatch, req.squad_overrides)

    # Accumulate per resolved target (an existing player id, or a create-new name).
    # Multiple rows for the same player merge — a later non-empty cell wins.
    targets: dict = {}          # key -> target dict
    note_set: set = set()
    skipped_rows = 0
    unresolved_squads: set = set()

    def target_key(pm, nm):
        if pm.get("player_id"):
            return ("existing", pm["player_id"])
        if pm.get("status") == "new":
            return ("new", nm)
        return None

    for r in req.rows:
        nm = row_name(r)
        if not nm:
            continue
        pm = pmatch.get(nm) or {}
        key = target_key(pm, nm)
        if key is None:
            skipped_rows += 1
            continue
        patch, notes = prof.row_profile(_row_values(r, req.mapping))
        for nt in notes:
            note_set.add(f"{nm}: {nt}")

        tgt = targets.get(key)
        if tgt is None:
            tgt = {"kind": key[0], "player_id": (key[1] if key[0] == "existing" else None),
                   "raw_name": nm, "patch": {}, "squad": ("none", None)}
            targets[key] = tgt
        tgt["patch"].update(patch)

        if squad_col:
            sq = str(r.get(squad_col, "")).strip()
            choice = _squad_choice(sq, smatch)
            if choice[0] != "none":
                tgt["squad"] = choice
                if choice[0] == "unresolved":
                    unresolved_squads.add(sq)

    # Build the preview (only the fields that actually change).
    preview = []
    for (kind, ident), tgt in targets.items():
        if kind == "existing":
            p = player_by_id.get(ident)
            current = _current_profile(p) if p else {}
            name = (p.display_name if p else tgt["raw_name"])
        else:
            current = {}
            name = tgt["raw_name"]

        proposed = {}
        for f, v in tgt["patch"].items():
            if v is not None and v != current.get(f):
                proposed[f] = v

        squad_kind, squad_val = tgt["squad"]
        squad_info = None
        if squad_kind == "team":
            if squad_val != current.get("squad_team_id"):
                squad_info = {"action": "change" if current.get("squad_team_id") else "set",
                              "team_id": squad_val, "name": team_name_by_id.get(squad_val, "(team)")}
        elif squad_kind == "new":
            squad_info = {"action": "new", "team_id": None, "name": squad_val}
        elif squad_kind == "unresolved":
            squad_info = {"action": "unresolved", "team_id": None, "name": squad_val}

        change_count = len(_display_changes(proposed)) + (1 if (squad_info and squad_info["action"] in ("set", "change", "new")) else 0)

        preview.append({
            "player_id": ident if kind == "existing" else None,
            "player_name": name,
            "raw_name": tgt["raw_name"],
            "new": kind == "new",
            "current": {k: current.get(k) for k in (*prof.PLAYER_FIELDS, "squad_team_id")},
            "current_squad_name": team_name_by_id.get(current.get("squad_team_id")) if current.get("squad_team_id") else None,
            "proposed": proposed,
            "squad": squad_info,
            "change_count": change_count,
        })
    # Most-changed first, new players after existing ones with the same count.
    preview.sort(key=lambda x: (x["change_count"], not x["new"]), reverse=True)

    # Warnings.
    warnings = []
    unresolved_names = [n for n, m in pmatch.items()
                        if not m.get("player_id") and m.get("status") not in ("skip", "new")]
    if unresolved_names:
        warnings.append(f"{len(unresolved_names)} name(s) need a match before import: "
                        + ", ".join(unresolved_names[:8]) + ("…" if len(unresolved_names) > 8 else ""))
    ambiguous = [n for n, m in pmatch.items() if m.get("status") == "ambiguous"]
    if ambiguous:
        warnings.append(f"{len(ambiguous)} name(s) match two players. Merge those first: "
                        + ", ".join(ambiguous[:5]))
    if unresolved_squads:
        warnings.append(f"{len(unresolved_squads)} squad name(s) don't match a team, they'll be left unset "
                        "unless you pick or create one: " + ", ".join(sorted(unresolved_squads)[:6]))
    bad_values = sorted(note_set)
    if bad_values:
        warnings.append(f"{len(bad_values)} cell(s) couldn't be read and were left unchanged "
                        "(see the notes below).")

    new_count = sum(1 for k in targets if k[0] == "new")
    matched_count = sum(1 for k in targets if k[0] == "existing")
    fields_to_change = sum(p["change_count"] for p in preview)

    return {
        "columns_used": {f: _col(req.mapping, f) for f in prof.ALL_FIELDS if _col(req.mapping, f)},
        "players": [{"raw_name": n, **m} for n, m in pmatch.items()],
        "squads": [{"raw_label": lb, **m} for lb, m in smatch.items()],
        "preview": preview,
        "notes": bad_values[:60],
        "warnings": warnings,
        "totals": {
            "rows": len(req.rows),
            "players_matched": matched_count,
            "players_new": new_count,
            "players_unresolved": len(unresolved_names),
            "rows_skipped": skipped_rows,
            "fields_to_change": fields_to_change,
            "players_changing": sum(1 for p in preview if p["change_count"] > 0),
        },
        "_targets": list(targets.values()),
    }


# Bowling action+type collapse to one logical "bowling" change for counting/display.
def _display_changes(proposed: dict) -> list:
    seen = []
    for f in proposed:
        if f in ("bowling_action", "bowling_type"):
            if "bowling" not in seen:
                seen.append("bowling")
        else:
            seen.append(f)
    return seen


# ── endpoints ─────────────────────────────────────────────────────────────────

@router.post("/preview")
async def preview_upload(
    file: UploadFile = File(...),
    current_user: User = Depends(require_cap(MANAGE_PLAYERS)),
    club: Organisation = Depends(get_current_club),
):
    """Parse an uploaded CSV/XLSX → headers, sample rows, auto column mapping.
    No DB writes; the client holds the parsed rows and posts them back."""
    data = await file.read()
    if not data:
        raise HTTPException(422, "Empty file.")
    if len(data) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, "File too large (max 8 MB).")
    parsed = ingest.parse_upload(file.filename or "upload.csv", data)
    if not parsed["headers"]:
        raise HTTPException(422, "Could not find a header row in the file.")
    mapping = prof.suggest_column_mapping(parsed["headers"])
    return {
        "filename": file.filename,
        "headers": parsed["headers"],
        "field_options": list(prof.ALL_FIELDS),
        "field_labels": prof.FIELD_LABELS,
        "mapping_suggestions": mapping,
        "row_count": len(parsed["rows"]),
        "sample_rows": parsed["rows"][:10],
        "rows": parsed["rows"],
    }


@router.post("/resolve")
async def resolve(
    req: ResolveRequest,
    current_user: User = Depends(require_cap(MANAGE_PLAYERS)),
    club: Organisation = Depends(get_current_club),
    db: AsyncSession = Depends(get_db),
):
    """Match names + squads and return the change preview. Idempotent — call
    again whenever the column mapping or a match is adjusted."""
    out = await _resolve(db, club.id, req)
    out.pop("_targets", None)
    return out


def _stored_name(raw: str) -> str:
    """A create-new name → the 'Last, First' shape the rest of the app sorts on.
    Already-comma'd names pass through; a single token stays as-is."""
    raw = (raw or "").strip()
    if "," in raw:
        return raw
    toks = raw.split()
    if len(toks) < 2:
        return raw
    return f"{toks[-1]}, {' '.join(toks[:-1])}"


@router.post("/commit")
async def commit(
    req: CommitRequest,
    current_user: User = Depends(require_cap(MANAGE_PLAYERS)),
    club: Organisation = Depends(get_current_club),
    db: AsyncSession = Depends(get_db),
):
    """Apply the previewed changes: update matched players' profile fields, create
    any new players / squads, and assign squads. Empty cells are never written, so
    a player keeps any value the sheet doesn't supply."""
    resolved = await _resolve(db, club.id, req)
    targets = resolved["_targets"]

    # Resolve every distinct create-new squad label to a team up-front (one team
    # per label, reusing an existing same-named team), so several players sharing
    # a new squad all land on the same row.
    teams = await _org_teams(db, club.id)
    team_by_norm = {prof._norm_squad(t.name): t for t in teams}
    new_team_for_label: dict = {}
    for tgt in targets:
        skind, sval = tgt["squad"]
        if skind == "new" and sval and sval not in new_team_for_label:
            existing = team_by_norm.get(prof._norm_squad(sval))
            if existing is not None:
                new_team_for_label[sval] = existing.id
            else:
                t = Team(id=uuid.uuid4(), organisation_id=club.id, name=sval.strip(), source="manual")
                db.add(t)
                await db.flush()
                team_by_norm[prof._norm_squad(sval)] = t
                new_team_for_label[sval] = t.id

    players_updated = 0
    players_created = 0
    squads_assigned = 0
    fields_written = 0

    for tgt in targets:
        if tgt["kind"] == "new":
            player = Player(id=uuid.uuid4(), name=_stored_name(tgt["raw_name"]), organisation_id=club.id)
            db.add(player)
            await db.flush()
            players_created += 1
        else:
            player = await db.get(Player, uuid.UUID(tgt["player_id"]))
            if not player or player.organisation_id != club.id:
                continue

        touched = False
        for f, v in tgt["patch"].items():
            if v is not None and getattr(player, f, None) != v:
                setattr(player, f, v)
                fields_written += 1
                touched = True

        # Squad assignment (mirrors update_player_profile so team_members stays in sync).
        skind, sval = tgt["squad"]
        new_team_id = None
        if skind == "team" and sval:
            new_team_id = uuid.UUID(sval)
        elif skind == "new" and sval in new_team_for_label:
            new_team_id = new_team_for_label[sval]
        if new_team_id is not None and player.squad_team_id != new_team_id:
            old_team_id = player.squad_team_id
            player.squad_team_id = new_team_id
            await sync_squad_membership(db, club.id, player.id, old_team_id, new_team_id, current_user.id)
            squads_assigned += 1
            touched = True

        # A row that marks someone inactive takes them out of their squad too,
        # same rule as the profile screen (routers/players.py). Read off the
        # sheet's own patch, not the stored status, so a row that says nothing
        # about status never moves a player who is already inactive.
        if tgt["patch"].get("status") == "inactive" and player.squad_team_id is not None:
            old_squad_id = player.squad_team_id
            player.squad_team_id = None
            await sync_squad_membership(db, club.id, player.id, old_squad_id, None, current_user.id)
            touched = True

        if touched and tgt["kind"] != "new":
            players_updated += 1

    await log_activity(
        db, org_id=club.id, user_id=current_user.id, action="player_profile_import",
        target_type="players", target_id=(req.filename or "csv"),
        details={"updated": players_updated, "created": players_created,
                 "fields_written": fields_written, "squads_assigned": squads_assigned,
                 "filename": req.filename},
    )
    await db.commit()

    return {
        "players_updated": players_updated,
        "players_created": players_created,
        "fields_written": fields_written,
        "squads_assigned": squads_assigned,
        "rows_skipped": resolved["totals"]["rows_skipped"],
    }


# One list of template columns, shared by the CSV and the Excel version so the
# two can't drift. The date of birth example is written ISO (2012-03-04) on
# purpose: 03/04 is a different month either side of the Pacific, and while
# the parser does pick one (day-first, like the clubs this serves), a template
# should not be teaching anyone to write the ambiguous form.
_TEMPLATE_HEADERS = [
    "Name", "Email", "Phone", "Squad", "Role", "Batting", "Bowling", "Gender",
    "Date of birth", "Opening batter", "Overseas", "Overseas country",
    "Status", "Show on website", "Fees", "Training",
]


def _template_examples(squad_names: list | None = None) -> list:
    squad_names = squad_names or []
    first = squad_names[0] if squad_names else "1st XI"
    second = squad_names[1] if len(squad_names) > 1 else "Women's 1st XI"
    return [
        ["Smith, John", "john.smith@example.com", "0412 345 678", first,
         "All Rounder", "Right handed", "Right-arm fast-medium", "Male",
         "1998-09-14", "No", "No", "", "Active", "Show", "Financial", "At training"],
        ["Patel, Anjali", "anjali.patel@example.com", "0423 456 789", second,
         "Batter", "Left handed", "", "Female",
         "2012-03-04", "Yes", "No", "", "Active", "Show", "", ""],
    ]


@router.get("/template.csv")
async def template(
    current_user: User = Depends(get_current_user),
    club: Organisation = Depends(get_current_club),
):
    """A starter template. The wizard maps any headers — this is just a hint at
    the columns it understands. Every column except Name is optional."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(_TEMPLATE_HEADERS)
    for row in _template_examples():
        w.writerow(row)
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="player_profiles_template.csv"'},
    )


@router.get("/template.xlsx")
async def template_xlsx(
    current_user: User = Depends(get_current_user),
    club: Organisation = Depends(get_current_club),
    db: AsyncSession = Depends(get_db),
):
    """An Excel version of the template with **dropdowns** on the fixed-choice
    columns (Role, Batting, Bowling, Gender) and the club's BetterSelect squads in
    the Squad column, so the optional fields are filled by picking rather than
    typing. A plain CSV can't carry dropdowns, so this is the Excel companion to
    template.csv. Dropdowns suggest but don't force — you can still type a value
    (e.g. a brand-new squad to create), and the importer normalises either way."""
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    # The club's selection-pool squads (BetterSelect teams) for the Squad dropdown.
    teams = (await db.execute(
        select(Team).where(Team.organisation_id == club.id, Team.is_active.is_(True))
    )).scalars().all()
    squad_names = [t.name for t in sorted(teams, key=lambda t: (t.sequence or 0, (t.name or "").lower()))]

    headers = list(_TEMPLATE_HEADERS)
    examples = _template_examples(squad_names)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Players"
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for row in examples:
        ws.append(row)
    ws.freeze_panes = "A2"
    for i, width in enumerate([22, 28, 16, 22, 20, 16, 22, 10,
                               14, 14, 11, 18, 11, 16, 14, 15], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    # Written as text, not as Excel dates: the sheet parser reads every cell as
    # a string, and a text cell survives the round trip exactly as typed
    # instead of arriving as a locale-formatted date nobody chose.
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=headers.index("Date of birth") + 1,
                            max_col=headers.index("Date of birth") + 1):
        for c in row:
            c.number_format = "@"

    # A hidden "Lists" sheet holds the dropdown values; the main sheet references
    # them by range (sidesteps the 255-char inline-list limit and keeps it tidy).
    lists = wb.create_sheet("Lists")
    header_col = {h: i + 1 for i, h in enumerate(headers)}
    dropdowns = {
        "Squad": squad_names,
        "Role": prof.ROLE_VALUES,
        "Batting": prof.BATTING_VALUES,
        "Bowling": prof.BOWLING_VALUES,
        "Gender": prof.GENDER_VALUES,
        "Opening batter": prof.YES_NO_VALUES,
        "Overseas": prof.YES_NO_VALUES,
        "Status": prof.STATUS_VALUES,
        "Show on website": prof.WEBSITE_VALUES,
        "Fees": prof.FEES_VALUES,
        "Training": prof.TRAINING_VALUES,
    }
    MAX_ROW = 1000
    list_col = 1
    for header, values in dropdowns.items():
        if not values:
            continue
        letter = get_column_letter(list_col)
        lists.cell(row=1, column=list_col, value=header)
        for ri, v in enumerate(values, start=2):
            lists.cell(row=ri, column=list_col, value=v)
        dv = DataValidation(
            type="list", formula1=f"Lists!${letter}$2:${letter}${len(values) + 1}", allow_blank=True,
        )
        dv.showErrorMessage = False  # suggest the list but still allow a typed value
        ws.add_data_validation(dv)
        pcol = get_column_letter(header_col[header])
        dv.add(f"{pcol}2:{pcol}{MAX_ROW}")
        list_col += 1
    lists.sheet_state = "hidden"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="player_profiles_template.xlsx"'},
    )
